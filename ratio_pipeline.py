# -*- coding: utf-8 -*-
"""Midasratio 鉄骨・木造断面検定パイプライン（MIDASratioplot_fig.m の逐語移植）.

元コード (msrc/privatetool_function/MIDAS/):
  ratio/MIDASratioplot_fig.m
      応力読み込み（2行/要素形式→3行展開含む）、荷重ケース区切りの取得、
      断面別最大検定値保存セルの設定、梁要素の断面検定ループ(約1016-1283行)、
      トラス要素の断面検定ループ(約1425-1563行) の鉄骨経路 (ele_material(2)==1 or 2)
  ratio/l_case_ratio_analysis.m  荷重ケース種別判定
  ratio/qup_case.m               割増オプション
  ratio/analysis_case.m          検定荷重ケース生成 (combine_stress_b/_t 含む)
  ratio/getbuck.m, ratio/getbuck_uni.m  座屈長さ
  ratio/sectionget.m, ratio/sectionget2.m, ratio/stress_arrange.m,
  ratio/check_t_ratio.m          幅厚比検討 (po/hinge/steel_po_rank.m を含む)
  materialget.m, material_E.m, column_beam_judge.m, doublecheck.m
  ratio/S/S_truss_ratio_analysis.m
  ratio/W/W_ratio_analysis.m, ratio/W/W_truss_ratio_analysis.m (w_check.py 経由)

GUIダイアログの引数化 (対話なしの既定値はダイアログの既定ボタンに一致):
  - l_case_ratio_analysis: questdlg 既定 'Yes' で決まる決定的分岐のみ実装。
    (荷重ケース数 1/2/3/5/7/9)。それ以外の対話必須分岐は NotImplementedError。
  - qup_case: qup=None → せん断割増なし(1.0)。正負加力は既定 'Yes' → pm_direction=2。
  - analysis_case: トラス長期考慮 tr_long は既定 'Yes' → 1。
    短期荷重ケース名 inputdlg は既定文字列 (長期名+短期名の連結) を採用。
  - トラス検定のボルト欠損 listdlg は既定「欠損なし」(=0)、応力倍率は既定 1.00。

スコープ (今回の移植範囲):
  鉄骨 (material 2列目 = 1: SN系 / 2: 冷間成形角形鋼管等) と
  木造 (material 2列目 = 10, W_ratio_analysis / W_truss_ratio_analysis) と
  RC (material 2列目 = 3, rc_check.RC_ratio_analysis: RC梁(中実角/*REBAR-BEAM
  配筋) と RC壁柱(中実角/rc_params['rcw']配筋)。RC柱・TAPERED梁等の未移植
  分岐は NotImplementedError) を検定する。
  PC(4)/SRC(5)/CFT(6)/杭(7)/RM(8) は検定せず beam_ratio 該当行を
  0 とし、CheckResult.skipped に material_type 付きで記録する
  （MATLAB版はそれぞれの *_ratio_analysis を呼ぶため数値互換ではない点に注意）。
  木造の長期許容応力度は wal_up (長期=1.0 / 中長期(積雪)=1.3,
  MIDASratioplot_fig.m 冒頭の questdlg 相当) で指定する。
  面材壁 (板要素, W_plate_ratio_analysis) は未移植 (今後対応)。

インデックス規約: util.find_index は 0-based/見つからない場合 -1
  (MATLAB版は 1-based/0)。本モジュールでは読み替えて移植している。

単位系: 応力 kN・m 系 / 断面 m 系 (mgt) → s_check 内部で N/mm2・mm 系へ変換。
"""

import math
import os

import numpy as np

from .util import find_index, doublecheck, pick_text, loadtxt_tolerant
from .mgt import (
    mgtopen_node, mgtopen_element, mgtopen_material, mgtopen_length,
    mgtopen_unit_2015, node_index_get, mgtopen_RCbeam, mgtopen_RCcolumn,
    mgtopen_SRCbeam, mgtopen_SRCcolumn,
    mgtopen_thickness, mgtopen_plate, mgtopen_wall, mgtopen_RCwall,
)
from .section import mgtopen_section
from .alst import ALST_S
from .secprops import BSB, BSR, BP, BBOX
from .w_check import (
    W_ratio_analysis, W_truss_ratio_analysis, w_al_match,
)
from .rc_check import (RC_ratio_analysis, rcw_input_row, RC4_YOSE_FIX,
                       RCSR_NMZ_FIX)
from .src_check import SRC_ratio_analysis, SRCcolumn_ALNM, SRCbeam_ALM
# 未検証エンジン(杭/PC/RM/CFT/板/壁)の使用検知 → 実行時注記
UNVERIFIED_NOTE = {'hit': set()}

from .s_check import (
    _num2str,
    S_ratio_analysis,
    TH_analysis, TH_analysis_text,
    TBOX_analysis, TBOX_analysis_text,
    TP_analysis, TP_analysis_text,
    TC_analysis, TC_analysis_text,
    TL_analysis, TL_analysis_text,
    TSB_analysis, TSB_analysis_text,
    TSR_analysis, TSR_analysis_text,
    TSR_SIGN_FIX,
)


# ===========================================================================
# 共通ヘルパー
# ===========================================================================

def _empty(x):
    """MATLAB isempty 相当."""
    if x is None:
        return True
    if isinstance(x, (list, tuple, str)):
        return len(x) == 0
    return np.asarray(x).size == 0


def doublecheck_nonsort(original):
    """analysis_case.m サブ関数 doublecheck_nonsort の逐語移植.

    一列のベクトルから連続重複のみを削る（並び順維持）。
    """
    original = np.asarray(original, dtype=float).ravel()
    if original.size == 0:
        return original
    judge = np.diff(original)
    remain = np.where(judge)[0] + 1
    remain = np.concatenate(([0], remain)).astype(int)
    return original[remain]


# ===========================================================================
# materialget.m / material_E.m / column_beam_judge.m
# ===========================================================================

def materialget(element_no, element_info):
    """要素番号(スカラまたはベクトル)から材料番号を返す (materialget.m)."""
    element_index = find_index(element_info[:, 0], element_no)
    return element_info[element_index, 1]


def material_E(material, material_no):
    """material_E.m の逐語移植 (S材のみ本移植の対象。RC系は E_RC_AIJ 未移植)."""
    mat_type = material[find_index(material[:, 0], material_no), 1]
    if mat_type == 1 or mat_type == 2:  # S
        return 205000 * 10 ** 3  # N/mm2→kN/m2
    elif mat_type == 10:  # W
        return 10  # （仮）木は10kN/m2とする
    raise NotImplementedError(
        'material_E: RC/PC/SRC (E_RC_AIJ) は移植スコープ外 (mat_type=%s)' % mat_type)


def column_beam_judge(element_index, element_info, node_info):
    """部材が水平要素（梁）か鉛直部材（柱）かを判定 (column_beam_judge.m).

    element_index は element_info の行index (0-based) のスカラまたはベクトル。
    柱=2 (完全鉛直), 梁=1, ブレースorななめ柱=1.9。
    """
    element_index = np.atleast_1d(np.asarray(element_index, dtype=int))
    num_element = element_index.size
    c_b_result = np.zeros(num_element)
    for i in range(num_element):
        # MATLAB: element_info(element_index(i)) は線形index→1列目の要素番号
        ele_no = element_info[element_index[i], 0]
        nodei_index, nodej_index = node_index_get(ele_no, element_info, node_info)
        XY = math.sqrt(
            (node_info[nodei_index, 1] - node_info[nodej_index, 1]) ** 2
            + (node_info[nodei_index, 2] - node_info[nodej_index, 2]) ** 2)
        Z = abs(node_info[nodei_index, 3] - node_info[nodej_index, 3])
        if XY > Z:  # 45度以上の傾きがあれば梁とする
            c_b_result[i] = 1
        elif XY < Z / 100:
            c_b_result[i] = 2  # 完全に鉛直な柱
        else:
            c_b_result[i] = 1.9  # ブレースorななめ柱
    return c_b_result


# ===========================================================================
# sectionget.m / sectionget2.m
# ===========================================================================

def sectionget(ele_no, element_info, section_info, limit_sec_no):
    """要素番号から断面情報行を取得する (sectionget.m).

    戻り値: 1次元 ndarray
      [断面番号, 寸法..., section_index1(=断面形式コード), section_index2]
      ダミー材 (断面番号>=limit_sec_no) は [断面番号 0 0 0]。
    """
    element_index = find_index(element_info[:, 0], ele_no)
    section_no = element_info[element_index, 2]
    if limit_sec_no <= section_no:
        return np.array([section_no, 0, 0, 0], dtype=float)

    section_index2 = 0  # MATLAB 1-based 累積 (見つからない場合の加算は0)
    section_index1 = None
    for i in range(len(section_info)):
        tbl = section_info[i]
        if _empty(tbl):
            continue
        fi = find_index(np.asarray(tbl)[:, 0], section_no)
        section_index2 += (fi + 1)  # MATLAB版 find_index 相当 (0=なし)
        if fi != -1:
            section_index1 = i
    tbl = np.asarray(section_info[section_index1], dtype=float)
    return np.concatenate((
        [section_no], tbl[section_index2 - 1, 1:tbl.shape[1]],
        [section_index1, section_index2]))


def sectionget2(section_info, section_no):
    """断面番号から断面情報行を取得する (sectionget2.m)."""
    section_index2 = 0
    section_index1 = None
    for i in range(len(section_info)):
        tbl = section_info[i]
        if _empty(tbl):
            continue
        fi = find_index(np.asarray(tbl)[:, 0], section_no)
        section_index2 += (fi + 1)
        if fi != -1:
            section_index1 = i
    tbl = np.asarray(section_info[section_index1], dtype=float)
    return np.concatenate((
        [section_no], tbl[section_index2 - 1, 1:tbl.shape[1]],
        [section_index1, section_index2]))


# ===========================================================================
# getbuck.m / getbuck_uni.m
# ===========================================================================

def _buck_one(ele_length, element_no, tab, scale_by_length):
    if _empty(tab):
        return ele_length
    tab = np.asarray(tab, dtype=float)
    idx = find_index(tab[:, 0], element_no)
    if idx == -1:
        return ele_length
    if scale_by_length:
        return ele_length * tab[idx, 1]
    return tab[idx, 1]


def getbuck(ele_length, element_no, lky, lkz, lb, select_length,
            cb_judge, H_beam_RCsupport, stress):
    """座屈長さ・圧縮フランジ間距離の設定 (getbuck.m).

    select_length: 1=実長指定, 2=部材長さの倍数指定, 3=lb以外を部材長とする。
    (圧縮フランジ間距離の応力による算定部は元コードでコメントアウト済のため未移植)
    """
    buck_length = np.zeros(3)
    if select_length == 1:  # 実長で座屈長さなど指定
        buck_length[0] = _buck_one(ele_length, element_no, lky, False)
        buck_length[1] = _buck_one(ele_length, element_no, lkz, False)
        buck_length[2] = _buck_one(ele_length, element_no, lb, False)
    elif select_length == 2:  # 部材の倍数で座屈長さなど指定
        buck_length[0] = _buck_one(ele_length, element_no, lky, True)
        buck_length[1] = _buck_one(ele_length, element_no, lkz, True)
        buck_length[2] = _buck_one(ele_length, element_no, lb, True)
    elif select_length == 3:
        buck_length[0] = ele_length
        buck_length[1] = ele_length
    else:
        raise RuntimeError('getbuck: select_length不正')  # MATLAB: stop
    return buck_length


def getbuck_uni(ele_length, element_no, lky, lkz, lb, select_length,
                cb_judge, H_beam_RCsupport, stress, timecase,
                unit_nodes, in_unit_element):
    """結合部材用の座屈長さ設定 (getbuck_uni.m)。本体は getbuck と同一."""
    return getbuck(ele_length, element_no, lky, lkz, lb, select_length,
                   cb_judge, H_beam_RCsupport, stress)


# ===========================================================================
# stress_arrange.m
# ===========================================================================

def stress_arrange(original_stress, load_case_num):
    """入力応力（鉛直時と水平時）から荷重ケース検定用応力を作成 (stress_arrange.m).

    load_case_num: 1=長期のみ 2=一方向正加力 3=一方向正負加力
                   4=2方向正加力 5=2方向正負加力
    """
    original_stress = np.asarray(original_stress, dtype=float)
    size2 = original_stress.shape[1]

    def _check_blocks(nblock, msg):
        load_num = doublecheck(original_stress[:, 1])
        if load_num.shape[0] != nblock:
            print(msg)
            raise RuntimeError(msg)  # MATLAB: stop
        num = original_stress.shape[0] // nblock
        for k in range(nblock):
            sub = doublecheck(original_stress[k * num:(k + 1) * num, 1])
            if sub.shape[0] != 1:
                print(msg)
                raise RuntimeError(msg)
        return num

    if load_case_num == 1:
        load_num = doublecheck(original_stress[:, 1])
        if load_num.shape[0] != 1:
            print('ERROR:長期のみ検定！！入力荷重数過多')
            raise RuntimeError('長期のみ検定！！入力荷重数過多')
        # 元コード: add_stress = original_stres (タイポ) → MATLABではエラー
        return original_stress

    elif load_case_num == 2:
        num = _check_blocks(2, 'ERROR:一方向正加力を検定！！入力荷重数ミス')
        part1 = original_stress[0:num, :].copy()
        part2 = original_stress[0:num, :].copy()
        part2[:, 1] = 21
        part2[:, 2:size2] = (original_stress[0:num, 2:size2]
                             + original_stress[num:2 * num, 2:size2])
        return np.vstack([part1, part2])

    elif load_case_num == 3:
        num = _check_blocks(2, 'ERROR:一方向正負加力を検定！！入力荷重数ミス')
        part1 = original_stress[0:num, :].copy()
        part2 = original_stress[0:num, :].copy()
        part3 = original_stress[0:num, :].copy()
        part2[:, 1] = 21
        part3[:, 1] = 23
        part2[:, 2:size2] = (original_stress[0:num, 2:size2]
                             + original_stress[num:2 * num, 2:size2])
        part3[:, 2:size2] = (original_stress[0:num, 2:size2]
                             - original_stress[num:2 * num, 2:size2])
        return np.vstack([part1, part2, part3])

    elif load_case_num == 4:
        num = _check_blocks(3, 'ERROR:2方向正加力を検定！！入力荷重数ミス')
        part1 = original_stress[0:num, :].copy()
        part2 = original_stress[0:num, :].copy()
        part3 = original_stress[0:num, :].copy()
        part2[:, 1] = 21
        part3[:, 1] = 22
        part2[:, 2:size2] = (original_stress[0:num, 2:size2]
                             + original_stress[num:2 * num, 2:size2])
        part3[:, 2:size2] = (original_stress[0:num, 2:size2]
                             + original_stress[2 * num:3 * num, 2:size2])
        return np.vstack([part1, part2, part3])

    elif load_case_num == 5:
        num = _check_blocks(3, 'ERROR:2方向正負加力を検定！！入力荷重数ミス')
        part1 = original_stress[0:num, :].copy()
        parts = [original_stress[0:num, :].copy() for _ in range(4)]
        for p, no in zip(parts, (21, 22, 23, 24)):
            p[:, 1] = no
        parts[0][:, 2:size2] = (original_stress[0:num, 2:size2]
                                + original_stress[num:2 * num, 2:size2])
        parts[1][:, 2:size2] = (original_stress[0:num, 2:size2]
                                + original_stress[2 * num:3 * num, 2:size2])
        parts[2][:, 2:size2] = (original_stress[0:num, 2:size2]
                                - original_stress[num:2 * num, 2:size2])
        parts[3][:, 2:size2] = (original_stress[0:num, 2:size2]
                                - original_stress[2 * num:3 * num, 2:size2])
        return np.vstack([part1] + parts)

    raise RuntimeError('stress_arrange: load_case_num不正')


# ===========================================================================
# steel_po_rank.m (check_t_ratio用)
# ===========================================================================

def steel_po_rank(sectionsize, c_b_index, SN, L):
    """鉄骨部材の保有水平耐力計算時 部材ランク算定 (steel_po_rank.m).

    戻り値: (steel_rank, ratio_f, ratio_w, ratio_t)
    steel_rank: 1=FA, 11=FAA, 2=FB, 3=FC, 4=FD, 0=対象外
    """
    sectionsize = np.asarray(sectionsize, dtype=float).ravel()
    ratio_f = None
    ratio_w = None
    ratio_t = None

    F = ALST_S(SN)
    F = F[2, 0]

    k_table = np.array([
        [22, 27, 32], [71, 87, 104], [71, 71, 74], [26, 33, 40],
        [63, 77, 94], [71, 71, 74], [22, 27, 32], [144, 175, 209],
        [100, 100, 110], [26, 33, 40], [118, 147, 180], [100, 100, 110]],
        dtype=float)

    n = sectionsize.size
    code = sectionsize[n - 2]

    H = B = t1 = t2 = None
    if c_b_index == 2 and code == 1:
        H, B, t1, t2 = sectionsize[0], sectionsize[1], sectionsize[2], sectionsize[3]
        if SN == 400:
            k_table = k_table[0:3, :]
        elif SN == 490:
            k_table = k_table[3:6, :]
    elif c_b_index == 1 and code == 1:
        H, B, t1, t2 = sectionsize[0], sectionsize[1], sectionsize[2], sectionsize[3]
        if SN == 400:
            k_table = k_table[6:9, :]
        elif SN == 490:
            k_table = k_table[9:12, :]

    def _faa_check():
        kf = k_table[0, 0]
        kw = k_table[1, 0]
        kc = k_table[2, 0]
        checkr1 = ((B / 2 / t2) ** 2 / (kf / math.sqrt(F / 98)) ** 2
                   + ((H - 2 * t2) / t1) ** 2 / (kw / math.sqrt(F / 98)) ** 2)
        checkr2 = (H - 2 * t2) / t1 / kc * math.sqrt(F / 98)
        return checkr1 <= 1 and checkr2 <= 1

    if c_b_index == 2:  # 柱のとき
        if code == 1:  # H鋼
            ratio_f = (sectionsize[1] - sectionsize[2]) / 2 / sectionsize[3]
            ratio_w = (sectionsize[0] - sectionsize[3] * 2) / sectionsize[2]
            if sectionsize[4] != 0:
                ratio_f2 = (sectionsize[4] - sectionsize[2]) / 2 / sectionsize[5]
                ratio_w2 = ((sectionsize[0] - sectionsize[3] - sectionsize[5])
                            / sectionsize[2])
                ratio_f = max(ratio_f, ratio_f2)
                ratio_w = max(ratio_w, ratio_w2)
            if 9.5 * math.sqrt(235 / F) > ratio_f and 43 * math.sqrt(235 / F) > ratio_w:
                steel_rank = 1
            elif 12 * math.sqrt(235 / F) > ratio_f and 45 * math.sqrt(235 / F) > ratio_w:
                steel_rank = 11 if _faa_check() else 2
            elif (15.5 * math.sqrt(235 / F) > ratio_f
                  and 48 * math.sqrt(235 / F) > ratio_w):
                steel_rank = 11 if _faa_check() else 3
            else:
                steel_rank = 4
        elif code == 2:  # 中実角 FA
            steel_rank = 1
        elif code == 4:  # 鋼管
            ratio_t = sectionsize[0] / sectionsize[1]
            if 50 * (235 / F) > ratio_t:
                steel_rank = 1
            elif 70 * (235 / F) > ratio_t:
                steel_rank = 2
            elif 100 * (235 / F) > ratio_t:
                steel_rank = 3
            else:
                steel_rank = 4
        elif code == 7:  # 角形鋼管
            ratio_t = max(sectionsize[0], sectionsize[1]) / sectionsize[2]
            if 33 * math.sqrt(235 / F) > ratio_t:
                steel_rank = 1
            elif 37 * math.sqrt(235 / F) > ratio_t:
                steel_rank = 2
            elif 48 * math.sqrt(235 / F) > ratio_t:
                steel_rank = 3
            else:
                steel_rank = 4
        else:
            steel_rank = 0

    elif c_b_index == 1:  # 梁のとき
        if code == 1:  # H鋼
            ratio_f = (sectionsize[1] - sectionsize[2]) / 2 / sectionsize[3]
            ratio_w = (sectionsize[0] - sectionsize[3] * 2) / sectionsize[2]
            if sectionsize[4] != 0:
                ratio_f2 = (sectionsize[4] - sectionsize[2]) / 2 / sectionsize[5]
                ratio_w2 = ((sectionsize[0] - sectionsize[3] - sectionsize[5])
                            / sectionsize[2])
                ratio_f = max(ratio_f, ratio_f2)
                ratio_w = max(ratio_w, ratio_w2)
            if 9 * math.sqrt(235 / F) > ratio_f and 60 * math.sqrt(235 / F) > ratio_w:
                steel_rank = 1
            elif 11 * math.sqrt(235 / F) > ratio_f and 65 * math.sqrt(235 / F) > ratio_w:
                steel_rank = 11 if _faa_check() else 2
            elif (15.5 * math.sqrt(235 / F) > ratio_f
                  and 71 * math.sqrt(235 / F) > ratio_w):
                steel_rank = 11 if _faa_check() else 3
            else:
                steel_rank = 4
        else:
            steel_rank = 0

    elif c_b_index == 1.9:  # ブレースのとき
        if code == 2:
            is_ = min(BSB(sectionsize[0:2] * 1000, 4), BSB(sectionsize[0:2] * 1000, 5))
        elif code == 3:
            is_ = BSR(sectionsize[0] * 1000, 4)
        elif code == 4:
            is_ = BP(sectionsize[0:2] * 1000, 4)
        elif code == 7:
            is_ = min(BBOX(sectionsize[0:4] * 1000, 4), BBOX(sectionsize[0:4] * 1000, 5))
        else:
            is_ = None
        if is_ is None:
            steel_rank = 0
        else:
            lamda = L * 10 ** 3 / (is_ * 10)
            if lamda >= 1980 / math.sqrt(F):
                steel_rank = 2  # BB
            elif lamda > 890 / math.sqrt(F):
                steel_rank = 3  # BC
            elif lamda > 495 / math.sqrt(F):
                steel_rank = 2  # BB
            else:
                steel_rank = 1  # BA
    else:
        steel_rank = 0

    return steel_rank, ratio_f, ratio_w, ratio_t


# ===========================================================================
# check_t_ratio.m
# ===========================================================================

_RANK_STR = {1: 'FA', 11: 'FAA', 2: 'FB', 3: 'FC', 4: 'FD'}


def check_t_ratio(sections, section_no, section_name):
    """使用断面の幅厚比検討・部材ランク表のTeX生成 (check_t_ratio.m).

    戻り値: list[str] (MATLAB版はstrvcatによるchar行列)
    """
    tex = ['\\section{柱梁の幅厚比の検討}']
    tex.append('使用断面の幅厚比の検討を行い使用材料，部位に対する部材ランクの算定を行う．')
    tex.append('\\def\\arraystretch{1.2}')
    tex.append('\\begin{center}')
    tex.append('\\footnotesize')
    tex.append('\\begin{longtable}{p{6zw}p{6zw}p{10zw}p{6zw}'
               '>{\\centering\\arraybackslash}p{4zw}>{\\centering\\arraybackslash}p{6zw}'
               '>{\\centering\\arraybackslash}p{6zw}>{\\centering\\arraybackslash}p{6zw}}')
    tex.append('\\caption{使用断面の幅厚比と材料区分，使用部位による部材ランク}')
    tex.append('\\label{lb_table}')
    tex.append('\\\\\\toprule')
    header = ('断面ID & 断面符号 & 断面サイズ & 使用箇所 & 幅厚比(1) & '
              '幅厚比(2)/$B/2t-f$ & 部材ランク(400) & 部材ランク(490) \\\\')
    tex.append(header)
    tex.append('\\bottomrule')
    tex.append('\\endfirsthead')
    tex.append('\\toprule')
    tex.append(header)
    tex.append('\\bottomrule')
    tex.append('\\endhead')

    section_no = np.atleast_1d(np.asarray(section_no, dtype=float).ravel())
    for i in range(section_no.shape[0]):
        sectionsize = sectionget2(sections, section_no[i])
        sectionsize = sectionsize[1:]

        sec_tex_no = _num2str(section_no[i], '%15.0f')
        sec_tex_name = pick_text(section_name[i], '_', -1)
        code = sectionsize[len(sectionsize) - 2]

        if code == 1:  # H鋼断面
            sec_tex_seize = ('H-' + _num2str(sectionsize[0] * 1000) + 'x'
                             + _num2str(sectionsize[1] * 1000) + 'x'
                             + _num2str(sectionsize[2] * 1000) + 'x'
                             + _num2str(sectionsize[3] * 1000))
            for c_b, place in ((2, '柱'), (1, '梁')):
                rank_400 = steel_po_rank(sectionsize, c_b, 400, 0)[0]
                rank_490, ratio_f, ratio_w, ratio_t = steel_po_rank(
                    sectionsize, c_b, 490, 0)
                t_ratio1 = _num2str(ratio_w, '%15.0f')
                t_ratio2 = _num2str(ratio_f, '%15.0f')
                tex.append(sec_tex_no + '&' + sec_tex_name + '&' + sec_tex_seize
                           + '&' + place + '&' + t_ratio1 + '&' + t_ratio2 + '&'
                           + _RANK_STR[rank_400] + '&' + _RANK_STR[rank_490] + '\\\\')

        elif code == 4:  # 丸鋼管
            sec_tex_seize = ('○-' + _num2str(sectionsize[0] * 1000) + 'x'
                             + _num2str(sectionsize[1] * 1000))
            rank_400 = steel_po_rank(sectionsize, 2, 400, 0)[0]
            rank_490, ratio_f, ratio_w, ratio_t = steel_po_rank(sectionsize, 2, 490, 0)
            t_ratio1 = _num2str(ratio_t, '%15.0f')
            t_ratio2 = _num2str(0, '%15.0f')
            tex.append(sec_tex_no + '&' + sec_tex_name + '&' + sec_tex_seize
                       + '&柱&' + t_ratio1 + '&' + t_ratio2 + '&'
                       + _RANK_STR[rank_400] + '&' + _RANK_STR[rank_490] + '\\\\')

        elif code == 7:  # 角鋼管
            sec_tex_seize = ('□-' + _num2str(sectionsize[0] * 1000) + 'x'
                             + _num2str(sectionsize[1] * 1000) + 'x'
                             + _num2str(sectionsize[2] * 1000) + 'x'
                             + _num2str(sectionsize[3] * 1000))
            rank_400 = steel_po_rank(sectionsize, 2, 400, 0)[0]
            rank_490, ratio_f, ratio_w, ratio_t = steel_po_rank(sectionsize, 2, 490, 0)
            t_ratio1 = _num2str(ratio_t, '%15.0f')
            t_ratio2 = _num2str(0, '%15.0f')
            tex.append(sec_tex_no + '&' + sec_tex_name + '&' + sec_tex_seize
                       + '&柱&' + t_ratio1 + '&' + t_ratio2 + '&'
                       + _RANK_STR[rank_400] + '&' + _RANK_STR[rank_490] + '\\\\')

    tex.append('\\hline')
    tex.append('\\end{longtable}')
    tex.append('\\end{center}')
    tex.append('以上から全ての断面において部材ランクがルート2の条件を満たしていることの確認が行えた．')
    tex.append('また表内でFAAとなっているものについては昭55建告第1791号第四号ただし書の規定を'
               '適用してSN材に対して検討を行った結果である．')
    return tex


# ===========================================================================
# l_case_ratio_analysis.m (非対話・既定回答による決定的分岐のみ)
# ===========================================================================

def l_case_ratio_analysis(num_load_case, load_case_no):
    """読み込まれた応力のタイプを判別する (l_case_ratio_analysis.m).

    GUI (questdlg/listdlg/inputdlg) は既定回答 (既定ボタン='Yes') に固定した
    非対話版。既定回答で決まらない分岐は NotImplementedError。

    戻り値: (L_Lcase, H_Lcase, S_Lcase, load_case_name, load_direction)
      L/H/S_Lcase は MATLAB版同様 1-based のケース順位 (ndarray)。
      load_case_name は list[str]。
      load_direction は ndarray (ケース数×3: [ケース番号, dx, dy]) または空。
    """
    load_case_no = np.asarray(load_case_no, dtype=float).ravel()
    if num_load_case != load_case_no.size:
        print('荷重ケース数と荷重ラベルが一致しません')
        raise RuntimeError('荷重ケース数と荷重ラベルが一致しません')  # MATLAB: stop

    L_Lcase = np.array([], dtype=int)
    H_Lcase = np.array([], dtype=int)
    S_Lcase = np.array([], dtype=int)
    load_case_name = []
    load_direction = np.zeros((0, 3))

    if num_load_case == 0:
        print('荷重ケースが存在しません！！')

    elif num_load_case == 1:
        # questdlg「読み込み荷重ケースは長期荷重ですか？」既定 'Yes'
        load_case_name = ['G+P']
        L_Lcase = np.array([1])

    elif num_load_case == 2:
        # questdlg「積雪検討（中長期と中短期）ですか？」既定 'Yes'
        L_Lcase = np.array([1])
        S_Lcase = np.array([2])
        load_case_name = ['中長期', '中短期']
        load_direction = np.array([[load_case_no[0], 0, 0],
                                   [load_case_no[1], 0, 0]])

    elif num_load_case == 3:
        # questdlg「長期+2方向（XY）の水平荷重ですか？」既定 'Yes'
        L_Lcase = np.array([1])
        H_Lcase = np.array([2, 3])
        load_case_name = ['TL', 'KX', 'KY']
        load_direction = np.array([[load_case_no[0], 0, 0],
                                   [load_case_no[1], 1, 0],
                                   [load_case_no[2], 0, 1]])

    elif num_load_case == 5:
        # questdlg「長期+2方向（XY）正負の水平荷重ですか？」既定 'Yes'
        L_Lcase = np.array([1])
        H_Lcase = np.array([2, 3, 4, 5])
        load_case_name = ['G+P', '+KX', '-KX', '+KY', '-KY']
        load_direction = np.array([[load_case_no[0], 0, 0],
                                   [load_case_no[1], 1, 0],
                                   [load_case_no[2], -1, 0],
                                   [load_case_no[3], 0, 1],
                                   [load_case_no[4], 0, -1]])

    elif num_load_case == 7:
        # questdlg「長期＋30度刻み方向ですか？」既定 'Yes' (松原仕様)
        L_Lcase = np.array([1])
        H_Lcase = np.array([2, 3, 4, 5, 6, 7])
        load_case_name = ['G+P', '025', '055', '085', '115', '145', '175']
        load_direction = np.array([
            [load_case_no[0], 0, 0], [load_case_no[1], 1, 0],
            [load_case_no[2], 0, 1], [load_case_no[3], 0.866, 0.5],
            [load_case_no[4], 0.5, 0.866], [load_case_no[5], -0.5, 0.866],
            [load_case_no[6], -0.866, 0.5]])

    elif num_load_case == 9:
        # questdlg「べにっこ仕様（長期＋短期8ケース）ですか？」既定 'Yes' (つねいし)
        L_Lcase = np.array([1])
        S_Lcase = np.array([2, 3, 4, 5, 6, 7, 8, 9])
        load_case_name = ['G+P     ', 'G+P+KX  ', 'G+P+KY  ', 'G+P+K30 ',
                          'G+P+K330', 'G+P-KX  ', 'G+P-KY  ', 'G+P+K210',
                          'G+P+K150']
        load_direction = np.array([
            [load_case_no[0], 0, 0], [load_case_no[1], 1, 0],
            [load_case_no[2], -1, 0], [load_case_no[3], 0, 1],
            [load_case_no[4], 0, -1], [load_case_no[5], 1, 1],
            [load_case_no[6], -1, 1], [load_case_no[7], -1, -1],
            [load_case_no[8], 1, -1]])

    else:
        raise NotImplementedError(
            'l_case_ratio_analysis: 荷重ケース数 %d は対話選択が必要なため'
            '非対話版では未対応' % num_load_case)

    # 全てのケースを判別できているかチェック
    if (int(np.sum(L_Lcase) + np.sum(H_Lcase) + np.sum(S_Lcase))
            != int(np.sum(np.arange(1, num_load_case + 1)))):
        print([L_Lcase, H_Lcase, S_Lcase])
        print(load_case_name)
        raise RuntimeError('荷重ケース判別エラー')  # MATLAB: stop

    # 水平荷重と短期荷重ケースが混じっていないかチェック
    if not _empty(H_Lcase) and not _empty(S_Lcase):
        print([L_Lcase, H_Lcase, S_Lcase])
        print(load_case_name)
        raise RuntimeError('水平荷重と短期荷重が混在')  # MATLAB: stop

    return L_Lcase, H_Lcase, S_Lcase, load_case_name, load_direction


# ===========================================================================
# ケース種別の明示指定 (l_case_ratio_analysis の対話選択の引数化)
# ===========================================================================

def _direction_from_name(name):
    """ケース名から荷重方向 (dx, dy) を推定する (KX/KY表記のみ判定).

    MATLAB l_case_ratio_analysis の listdlg「横力種類」選択に対応する
    load_direction 値: KX→(1,0) / KY→(0,1) / 負号付きは反転 / その他→(0,0)。
    """
    s = str(name).upper().replace(' ', '')
    if '-KX' in s:
        return (-1.0, 0.0)
    if 'KX' in s:
        return (1.0, 0.0)
    if '-KY' in s:
        return (0.0, -1.0)
    if 'KY' in s:
        return (0.0, 1.0)
    return (0.0, 0.0)


def case_types_to_lcase(num_load_case, load_case_no, case_types):
    """明示指定されたケース種別から L/H/S_Lcase を構成する.

    l_case_ratio_analysis の対話選択 (「長期荷重か？」「水平荷重のみ/
    短期荷重か？」「横力種類は？」) をユーザー指定で置き換えるもの。
    MATLAB原典と同じ形式の分類 (L_Lcase/H_Lcase/S_Lcase) を返し、
    後段の qup_case / analysis_case は原典どおりに動作する。

    case_types: {ケース番号: (種別, ケース名)} または {ケース番号: 種別}
      種別 'L': 長期荷重
      種別 'H': 水平荷重のみ → 長期と組合せて TL±H の検定ケースを生成
      種別 'S': 短期荷重（長期鉛直＋水平荷重の組合せ済み応力）→
                組合せ生成せずそのまま短期として検定 (長期は二重加算しない)。
                引張専用材等の収束計算済み応力はこれを選ぶ。
      ケース名省略時: L→'G+P', H→'CASE<番号>', S→'G+P+CASE<番号>'。

    MATLAB原典の timecase 規約 (>=10:短期 / 1,9:長期 / 他:stop) 上、
    'S' ケースのケース番号は10以上、'S' と併用する 'L' ケースは
    ケース番号1かつ先頭ブロックが必要 (冷間成形割増の長期応力参照も
    先頭ブロックを用いる原典仕様のため)。満たさない場合はエラー。

    戻り値: l_case_ratio_analysis と同一
      (L_Lcase, H_Lcase, S_Lcase, load_case_name, load_direction)
    """
    load_case_no = np.asarray(load_case_no, dtype=float).ravel()
    if num_load_case != load_case_no.size:
        print('荷重ケース数と荷重ラベルが一致しません')
        raise RuntimeError('荷重ケース数と荷重ラベルが一致しません')

    norm = {}
    for k, v in dict(case_types).items():
        norm[float(k)] = v
    unknown = [k for k in norm
               if not np.any(np.abs(load_case_no - k) == 0)]
    if unknown:
        print('注意: 指定されたケース %s は応力ファイルに存在しないため'
              '無視します' % sorted(unknown))

    L_list = []
    H_list = []
    S_list = []
    load_case_name = []
    ld = []
    missing = []
    for pos in range(num_load_case):
        cno = float(load_case_no[pos])
        if cno not in norm:
            missing.append(cno)
            continue
        v = norm[cno]
        if isinstance(v, (tuple, list)):
            ctype = str(v[0]).strip().upper()
            cname = str(v[1]).strip() if len(v) > 1 else ''
        else:
            ctype = str(v).strip().upper()
            cname = ''
        if ctype not in ('L', 'H', 'S'):
            raise RuntimeError(
                "ケース%gの種別 '%s' が不正です ('L'=長期/'H'=水平のみ/"
                "'S'=短期(組合せ済み))" % (cno, ctype))
        if not cname:
            if ctype == 'L':
                cname = 'G+P'
            elif ctype == 'H':
                cname = 'CASE%g' % cno
            else:
                cname = 'G+P+CASE%g' % cno
        if ctype == 'L':
            L_list.append(pos + 1)
        elif ctype == 'H':
            H_list.append(pos + 1)
        else:
            S_list.append(pos + 1)
        load_case_name.append(cname)
        dx, dy = _direction_from_name(cname)
        ld.append([cno, dx, dy])
    if missing:
        raise RuntimeError('ケース種別が未指定の荷重ケースがあります: %s'
                           % ['%g' % c for c in missing])

    # 水平荷重と短期荷重ケースが混じっていないかチェック (MATLAB原典と同一)
    if L_list and not H_list and not S_list:
        pass
    if H_list and S_list:
        print([L_list, H_list, S_list])
        print(load_case_name)
        raise RuntimeError("水平荷重('H')と短期荷重('S')は混在できません "
                           '(MATLAB原典 l_case_ratio_analysis と同一の制約)')

    # 'S'(組合せ済み)経路は応力ブロックをそのまま検定するため、
    # MATLAB原典の timecase 規約・長期応力参照の前提を満たす必要がある
    if S_list:
        bad = [load_case_no[p - 1] for p in S_list if load_case_no[p - 1] < 10]
        if bad:
            raise RuntimeError(
                "短期(組合せ済み)'S'指定ケースのケース番号は10以上が必要です"
                ' (MATLAB原典のtimecase規約: >=10を短期と判別): %s'
                % ['%g' % c for c in bad])
        if L_list and (L_list[0] != 1 or load_case_no[L_list[0] - 1] != 1):
            raise RuntimeError(
                "'S'指定と併用する長期'L'ケースはケース番号1かつ先頭ブロック"
                'が必要です (MATLAB原典は長期の判別(timecase==1)と冷間成形'
                '割増の長期応力参照に先頭ブロック=ケース1を用いるため)')

    return (np.array(L_list, dtype=int), np.array(H_list, dtype=int),
            np.array(S_list, dtype=int), load_case_name,
            np.asarray(ld, dtype=float) if ld else np.zeros((0, 3)))


def default_case_types(load_case_no):
    """自動判定 (l_case_ratio_analysis) による既定のケース種別一覧を返す.

    UI「ケース読込」の初期値用。
    戻り値: [{'no': ケース番号, 'type': 'L'/'H'/'S', 'name': ケース名}]
    自動判定が対話必須のケース数 (4,6,8,10以上) では先頭を 'L' (G+P)、
    残りを 'H' とする。
    """
    load_case_no = np.asarray(load_case_no, dtype=float).ravel()
    n = int(load_case_no.size)
    try:
        L_Lcase, H_Lcase, S_Lcase, names, _ld = l_case_ratio_analysis(
            n, load_case_no)
        L_set = set(int(v) for v in np.atleast_1d(L_Lcase))
        H_set = set(int(v) for v in np.atleast_1d(H_Lcase))
        out = []
        for pos in range(n):
            if (pos + 1) in L_set:
                t = 'L'
            elif (pos + 1) in H_set:
                t = 'H'
            else:
                t = 'S'
            out.append({'no': float(load_case_no[pos]), 'type': t,
                        'name': str(names[pos]).strip()})
        return out
    except (NotImplementedError, RuntimeError):
        out = []
        for pos in range(n):
            if pos == 0:
                out.append({'no': float(load_case_no[pos]), 'type': 'L',
                            'name': 'G+P'})
            else:
                out.append({'no': float(load_case_no[pos]), 'type': 'H',
                            'name': 'CASE%g' % load_case_no[pos]})
        return out


# ===========================================================================
# qup_case.m (GUI引数化)
# ===========================================================================

def qup_case(L_Lcase, H_Lcase, S_Lcase, qup=None, pm_direction=None):
    """割増オプションの決定 (qup_case.m のGUI引数化版).

    qup: None → せん断力の割増なし (questdlg 'No' 相当, 全て1.0)。
         dict {'beam':x,'wall':y,'src':z} → 数値指定 (listdlg 4番相当)。
    pm_direction: None → 該当分岐の questdlg 既定 'Yes' → 2 (正負2方向)。
                  1/2 指定で上書き。
    戻り値: (qup_beam, qup_wall, qup_beam_src, pm_direction)
    """
    def _qup_values():
        if qup is None:
            return 1.0, 1.0, 1.0
        return (float(qup.get('beam', 1.0)), float(qup.get('wall', 1.0)),
                float(qup.get('src', 1.0)))

    if _empty(H_Lcase) and _empty(S_Lcase):  # 長期荷重のみ
        return 1.0, 1.0, 1.0, 1

    elif _empty(L_Lcase) and _empty(H_Lcase):  # 短期荷重のみ
        return 1.0, 1.0, 1.0, 1

    elif _empty(L_Lcase) and _empty(S_Lcase):  # 水平荷重のみ
        qb, qw, qs = _qup_values()
        return qb, qw, qs, 1

    elif _empty(S_Lcase):  # 長期荷重ケースと水平荷重
        qb, qw, qs = _qup_values()
        pm = 2 if pm_direction is None else int(pm_direction)
        return qb, qw, qs, pm

    elif _empty(H_Lcase):  # 長期荷重と短期荷重
        qb, qw, qs = _qup_values()
        pm = 2 if pm_direction is None else int(pm_direction)
        return qb, qw, qs, pm

    else:  # H(水平)とS(短期)の両方が存在 (上流の種別チェックで弾かれる想定)
        raise ValueError(
            '荷重ケース種別の組合せが不正です: H(水平)とS(短期組合せ済み)の'
            '混在には対応していません。ケース種別指定を確認してください。')


# ===========================================================================
# analysis_case.m (板要素・壁要素は本移植ではスコープ外のため None 固定)
# ===========================================================================

def _same_load_case(num_load_case, load_case_no_ele):
    """analysis_case.m サブ関数 same_load_case."""
    if _empty(load_case_no_ele):
        return
    if num_load_case != np.asarray(load_case_no_ele).shape[0]:
        print('要素の荷重ケース数が全体のケースと一致しません．')
        print(num_load_case)
        print(load_case_no_ele)
        raise RuntimeError('要素の荷重ケース数不一致')  # MATLAB: stop


def _combine_stress_b(pm_direction, load_case_no, beam_stress,
                      load_case_no_beam, load_case_index_beam,
                      L_Lcase, H_Lcase):
    """analysis_case.m サブ関数 combine_stress_b の逐語移植.

    長期応力に水平応力を±合成して検定荷重ケースを作る。
    合成後ケース番号 = 長期ケース番号*10*i_pm + 水平ケース番号。
    ※せん断力の割増は元コードで *1.0 に固定されている (qup_beam未使用)。
    """
    nb = load_case_no_beam.shape[0]
    L_index = find_index(load_case_no_beam, load_case_no[L_Lcase[0] - 1])
    if L_index == nb - 1:
        L_stress = beam_stress[int(load_case_index_beam[L_index]):, :].copy()
    else:
        L_stress = beam_stress[int(load_case_index_beam[L_index]):
                               int(load_case_index_beam[L_index + 1]), :].copy()
    ch_beam_stress = L_stress.copy()
    ch_beam_stress[:, 1] = 1

    H_index = find_index(load_case_no_beam, load_case_no[np.asarray(H_Lcase) - 1])
    H_index = np.atleast_1d(H_index)
    H_stress = []
    for i_h in range(len(H_Lcase)):
        if H_index[i_h] == nb - 1:
            H_stress.append(beam_stress[int(load_case_index_beam[H_index[i_h]]):, :])
        else:
            H_stress.append(beam_stress[int(load_case_index_beam[H_index[i_h]]):
                                        int(load_case_index_beam[H_index[i_h] + 1]), :])

    for i_h in range(len(H_Lcase)):
        for i_pm in range(1, pm_direction + 1):
            sub_h_stress = H_stress[i_h].copy()
            sub_h_stress[:, 1] = (beam_stress[int(load_case_index_beam[L_index]), 1]
                                  * 10 * i_pm
                                  + beam_stress[int(load_case_index_beam[H_index[i_h]]), 1])
            if i_pm == 1:
                sub_h_stress[:, 3:5] = sub_h_stress[:, 3:5] * 1.0
                sub_h_stress[:, 2:8] = L_stress[:, 2:8] + sub_h_stress[:, 2:8]
            elif i_pm == 2:
                sub_h_stress[:, 3:5] = sub_h_stress[:, 3:5] * 1.0
                sub_h_stress[:, 2:8] = L_stress[:, 2:8] - sub_h_stress[:, 2:8]
            ch_beam_stress = np.vstack([ch_beam_stress, sub_h_stress])

    ch_load_case_no_beam = doublecheck_nonsort(ch_beam_stress[:, 1])
    ch_load_case_index_beam = find_index(ch_beam_stress[:, 1], ch_load_case_no_beam)
    return ch_beam_stress, ch_load_case_no_beam, ch_load_case_index_beam


def _combine_stress_t(pm_direction, load_case_no, truss_stress,
                      load_case_no_truss, load_case_index_truss,
                      L_Lcase, H_Lcase, tr_long):
    """analysis_case.m サブ関数 combine_stress_t の逐語移植."""
    nt = load_case_no_truss.shape[0]
    L_index = find_index(load_case_no_truss, load_case_no[L_Lcase[0] - 1])
    if L_index == nt - 1:
        L_stress = truss_stress[int(load_case_index_truss[L_index]):, :].copy()
    else:
        L_stress = truss_stress[int(load_case_index_truss[L_index]):
                                int(load_case_index_truss[L_index + 1]), :].copy()
    L_stress[:, 2:4] = L_stress[:, 2:4] * tr_long
    ch_truss_stress = L_stress.copy()
    ch_truss_stress[:, 1] = 1

    H_index = np.atleast_1d(
        find_index(load_case_no_truss, load_case_no[np.asarray(H_Lcase) - 1]))
    H_stress = []
    for i_h in range(len(H_Lcase)):
        if H_index[i_h] == nt - 1:
            H_stress.append(truss_stress[int(load_case_index_truss[H_index[i_h]]):, :])
        else:
            H_stress.append(truss_stress[int(load_case_index_truss[H_index[i_h]]):
                                         int(load_case_index_truss[H_index[i_h] + 1]), :])

    for i_h in range(len(H_Lcase)):
        for i_pm in range(1, pm_direction + 1):
            sub_h_stress = H_stress[i_h].copy()
            sub_h_stress[:, 1] = (truss_stress[int(load_case_index_truss[L_index]), 1]
                                  * 10 * i_pm
                                  + truss_stress[int(load_case_index_truss[H_index[i_h]]), 1])
            if i_pm == 1:
                sub_h_stress[:, 2:4] = L_stress[:, 2:4] + sub_h_stress[:, 2:4]
            elif i_pm == 2:
                sub_h_stress[:, 2:4] = L_stress[:, 2:4] - sub_h_stress[:, 2:4]
            ch_truss_stress = np.vstack([ch_truss_stress, sub_h_stress])

    ch_load_case_no_truss = doublecheck_nonsort(ch_truss_stress[:, 1])
    ch_load_case_index_truss = find_index(ch_truss_stress[:, 1], ch_load_case_no_truss)
    return ch_truss_stress, ch_load_case_no_truss, ch_load_case_index_truss


def analysis_case(L_Lcase, H_Lcase, S_Lcase, qup_beam, qup_wall, pm_direction,
                  load_case_no, beam_stress, load_case_no_beam,
                  load_case_index_beam, truss_stress, load_case_no_truss,
                  load_case_index_truss, load_case_name, num_load_case,
                  load_direction, tr_long=None,
                  wall_stress=None, load_case_no_wall=None,
                  load_case_index_wall=None, wall_stress_ID=None,
                  plate_stress=None, load_case_no_plate=None,
                  load_case_index_plate=None, plate_up=1.0, pl_long=None):
    """検定荷重ケースの作成 (analysis_case.m の逐語移植).

    tr_long: None → questdlg 既定 'Yes' → 1 (トラス検定で長期を考慮)。
    pl_long: None → questdlg 既定 'Yes' → 1 (板要素検定で長期を考慮)。
    plate_up: 板要素の水平応力割増係数 (inputdlg 既定 1.0)。
    戻り値: (load_case_no, ch_beam_stress, ch_load_case_no_beam,
             ch_load_case_index_beam, ch_truss_stress, ch_load_case_no_truss,
             ch_load_case_index_truss, LCNAME, check_case_num, load_direction2,
             ch_wall_stress, ch_load_case_no_wall, ch_load_case_index_wall,
             ch_wall_stress_ID, ch_plate_stress, ch_load_case_no_plate,
             ch_load_case_index_plate)
    """
    # トラス要素の検定において長期荷重を考慮するか
    if _empty(truss_stress):
        tr_long = 0
    else:
        if tr_long is None:
            tr_long = 1  # questdlg 既定 'Yes'
    # 板要素の検定において長期荷重を考慮するか
    if _empty(plate_stress):
        pl_long = 0
    else:
        if pl_long is None:
            pl_long = 1  # questdlg 既定 'Yes'
    load_direction2 = np.zeros((0, 2))

    # 空の既定
    ch_wall_stress = None
    ch_load_case_no_wall = None
    ch_load_case_index_wall = None
    ch_wall_stress_ID = None
    ch_plate_stress = None
    ch_load_case_no_plate = None
    ch_load_case_index_plate = None

    # 応力荷重ケース数が全体と一致しているかチェック
    _same_load_case(num_load_case, load_case_no_beam)
    _same_load_case(num_load_case, load_case_no_truss)
    _same_load_case(num_load_case, load_case_no_wall)
    _same_load_case(num_load_case, load_case_no_plate)

    # 総検討荷重ケース数の算出
    check_case_num = None
    if pm_direction == 2:
        if _empty(S_Lcase):
            check_case_num = (len(L_Lcase)
                              + len(L_Lcase) * 2 * len(H_Lcase))
        elif _empty(H_Lcase):
            check_case_num = (len(L_Lcase)
                              + len(L_Lcase) * 2 * len(S_Lcase))
    else:
        if _empty(S_Lcase) and _empty(H_Lcase):
            check_case_num = len(L_Lcase)
        elif _empty(S_Lcase) and _empty(L_Lcase):
            check_case_num = len(H_Lcase)
        elif _empty(S_Lcase):
            check_case_num = len(L_Lcase) + len(L_Lcase) * len(H_Lcase)
        elif _empty(H_Lcase) and _empty(L_Lcase):
            check_case_num = len(S_Lcase)
        elif _empty(H_Lcase):
            check_case_num = len(L_Lcase) + len(L_Lcase) * len(S_Lcase)
    if check_case_num is None:
        raise RuntimeError('analysis_case: check_case_num が定義できません'
                           ' (MATLAB版も未定義エラーとなる組合せ)')

    # 荷重ケース名の設定
    LCNAME = []
    if _empty(H_Lcase) and _empty(S_Lcase):  # 長期荷重のみ
        LCNAME = list(load_case_name)
        load_direction2 = load_direction
    elif _empty(L_Lcase) and _empty(H_Lcase):  # 短期荷重のみ
        LCNAME = list(load_case_name)
        load_direction2 = load_direction
    elif _empty(L_Lcase) and _empty(S_Lcase):  # 水平荷重のみ
        LCNAME = list(load_case_name)
        load_direction2 = load_direction
    elif _empty(S_Lcase):  # 長期荷重ケースと水平荷重
        ld2 = []
        for i_l in range(len(L_Lcase)):
            for i in range(1, check_case_num + 1):
                if i == 1:
                    LCNAME.append(load_case_name[L_Lcase[i_l] - 1])
                    ld2.append([0, 0])
                else:
                    # inputdlg の既定文字列をそのまま採用 (非対話)
                    if pm_direction == 2:
                        if i % 2 == 0:
                            s_name = (load_case_name[L_Lcase[i_l] - 1] + '+'
                                      + load_case_name[H_Lcase[i // 2 - 1] - 1])
                            ld2.append(list(
                                load_direction[H_Lcase[i // 2 - 1] - 1, 1:3]))
                        else:
                            s_name = (load_case_name[L_Lcase[i_l] - 1] + '-'
                                      + load_case_name[H_Lcase[i // 2 - 1] - 1])
                            ld2.append(list(
                                -1 * load_direction[H_Lcase[i // 2 - 1] - 1, 1:3]))
                    else:
                        s_name = (load_case_name[L_Lcase[i_l] - 1] + '+'
                                  + load_case_name[H_Lcase[i - 2] - 1])
                        ld2.append(list(load_direction[H_Lcase[i - 2] - 1, 1:3]))
                    LCNAME.append(s_name)
        load_direction2 = np.asarray(ld2, dtype=float)
    elif _empty(H_Lcase):  # 長期荷重と短期荷重 (すでに設定されている)
        LCNAME = list(load_case_name)
        load_direction2 = load_direction

    # 荷重ケースの作成
    if _empty(H_Lcase) and _empty(S_Lcase):  # 長期荷重のみ
        ch_beam_stress = None if _empty(beam_stress) else beam_stress
        ch_load_case_no_beam = None if _empty(beam_stress) else load_case_no_beam
        ch_load_case_index_beam = None if _empty(beam_stress) else load_case_index_beam
        ch_truss_stress = None if _empty(truss_stress) else truss_stress
        ch_load_case_no_truss = None if _empty(truss_stress) else load_case_no_truss
        ch_load_case_index_truss = (None if _empty(truss_stress)
                                    else load_case_index_truss)
        if not _empty(wall_stress):
            ch_wall_stress = wall_stress
            ch_load_case_no_wall = load_case_no_wall
            ch_load_case_index_wall = load_case_index_wall
            ch_wall_stress_ID = wall_stress_ID
        if not _empty(plate_stress):
            ch_plate_stress = plate_stress
            ch_load_case_no_plate = load_case_no_plate
            ch_load_case_index_plate = load_case_index_plate
        load_direction2 = load_direction

    elif _empty(L_Lcase) and _empty(H_Lcase):  # 短期荷重のみ
        ch_beam_stress = None if _empty(beam_stress) else beam_stress
        ch_load_case_no_beam = None if _empty(beam_stress) else load_case_no_beam
        ch_load_case_index_beam = None if _empty(beam_stress) else load_case_index_beam
        ch_truss_stress = None if _empty(truss_stress) else truss_stress
        ch_load_case_no_truss = None if _empty(truss_stress) else load_case_no_truss
        ch_load_case_index_truss = (None if _empty(truss_stress)
                                    else load_case_index_truss)
        if not _empty(wall_stress):
            ch_wall_stress = wall_stress
            ch_load_case_no_wall = load_case_no_wall
            ch_load_case_index_wall = load_case_index_wall
            ch_wall_stress_ID = wall_stress_ID
        if not _empty(plate_stress):
            ch_plate_stress = plate_stress
            ch_load_case_no_plate = load_case_no_plate
            ch_load_case_index_plate = load_case_index_plate
        load_direction2 = load_direction

    elif _empty(L_Lcase) and _empty(S_Lcase):  # 水平荷重のみ
        # 元コード: 「作成すること」(未実装) → 変数未定義でMATLABエラー
        raise NotImplementedError('analysis_case: 水平荷重のみは元コードで未実装')

    elif _empty(S_Lcase):  # 長期荷重ケースと水平荷重
        if _empty(beam_stress):
            ch_beam_stress = None
            ch_load_case_no_beam = None
            ch_load_case_index_beam = None
        else:
            (ch_beam_stress, ch_load_case_no_beam,
             ch_load_case_index_beam) = _combine_stress_b(
                pm_direction, load_case_no, beam_stress, load_case_no_beam,
                load_case_index_beam, L_Lcase, H_Lcase)
        if _empty(truss_stress):
            ch_truss_stress = None
            ch_load_case_no_truss = None
            ch_load_case_index_truss = None
        else:
            (ch_truss_stress, ch_load_case_no_truss,
             ch_load_case_index_truss) = _combine_stress_t(
                pm_direction, load_case_no, truss_stress, load_case_no_truss,
                load_case_index_truss, L_Lcase, H_Lcase, tr_long)
        if not _empty(wall_stress):
            (ch_wall_stress, ch_load_case_no_wall, ch_load_case_index_wall,
             ch_wall_stress_ID) = _combine_stress_w(
                pm_direction, load_case_no, wall_stress, load_case_no_wall,
                load_case_index_wall, wall_stress_ID, L_Lcase, H_Lcase,
                qup_wall)
        if not _empty(plate_stress):
            (ch_plate_stress, ch_load_case_no_plate,
             ch_load_case_index_plate) = _combine_stress_p(
                pm_direction, load_case_no, plate_stress,
                load_case_no_plate, load_case_index_plate, L_Lcase, H_Lcase,
                plate_up, pl_long)

    elif _empty(H_Lcase):  # 長期荷重と短期荷重
        load_direction2 = load_direction
        ch_beam_stress = None if _empty(beam_stress) else beam_stress
        ch_load_case_no_beam = None if _empty(beam_stress) else load_case_no_beam
        ch_load_case_index_beam = None if _empty(beam_stress) else load_case_index_beam
        ch_truss_stress = None if _empty(truss_stress) else truss_stress
        ch_load_case_no_truss = None if _empty(truss_stress) else load_case_no_truss
        ch_load_case_index_truss = (None if _empty(truss_stress)
                                    else load_case_index_truss)
        if not _empty(wall_stress):
            ch_wall_stress = wall_stress
            ch_load_case_no_wall = load_case_no_wall
            ch_load_case_index_wall = load_case_index_wall
            ch_wall_stress_ID = wall_stress_ID
        if not _empty(plate_stress):
            ch_plate_stress = plate_stress
            ch_load_case_no_plate = load_case_no_plate
            ch_load_case_index_plate = load_case_index_plate

    return (load_case_no, ch_beam_stress, ch_load_case_no_beam,
            ch_load_case_index_beam, ch_truss_stress, ch_load_case_no_truss,
            ch_load_case_index_truss, LCNAME, check_case_num, load_direction2,
            ch_wall_stress, ch_load_case_no_wall, ch_load_case_index_wall,
            ch_wall_stress_ID, ch_plate_stress, ch_load_case_no_plate,
            ch_load_case_index_plate)


def _doublecheck_nonsort(original):
    """analysis_case.m サブ関数 doublecheck_nonsort の逐語移植 (連続重複の除去)."""
    original = np.asarray(original, dtype=float).ravel()
    if original.size == 0:
        return original
    judge = np.diff(original)
    remain = np.where(judge != 0)[0] + 1
    remain = np.concatenate([[0], remain])
    return original[remain.astype(int)]


def _combine_stress_w(pm_direction, load_case_no, wall_stress,
                      load_case_no_wall, load_case_index_wall,
                      wall_stress_ID, L_Lcase, H_Lcase, qup_wall):
    """analysis_case.m サブ関数 combine_stress_w の逐語移植.

    壁要素は1データ=2行(上下)。水平ケースのせん断(列2:3)を qup_wall 倍して
    長期と±合成する。合成ケース番号は 長期ID*10*i_pm + 水平ID。
    """
    wall_stress = np.atleast_2d(np.asarray(wall_stress, dtype=float))
    wall_stress_ID = np.atleast_2d(np.asarray(wall_stress_ID, dtype=float))
    load_case_no_wall = np.asarray(load_case_no_wall, dtype=float).ravel()
    load_case_index_wall = np.asarray(load_case_index_wall).ravel()

    # 長期応力部分 (MATLAB 1-based index を保ったまま計算し、スライスで0-based化)
    L_index = int(find_index(load_case_no_wall,
                             load_case_no[int(L_Lcase[0]) - 1])) + 1
    if L_index == load_case_no_wall.size:
        L_stress = wall_stress[(load_case_index_wall[L_index - 1] + 1) * 2 - 1 - 1:, :].copy()
        L_ID = wall_stress_ID[load_case_index_wall[L_index - 1]:, :].copy()
    else:
        L_stress = wall_stress[(load_case_index_wall[L_index - 1] + 1) * 2 - 1 - 1:
                               (load_case_index_wall[L_index] + 1) * 2 - 2, :].copy()
        L_ID = wall_stress_ID[load_case_index_wall[L_index - 1]:
                              load_case_index_wall[L_index], :].copy()
    ch_wall_stress = L_stress.copy()
    # NOTE: 原典359行 ch_wall_stress(:,2)=1 は応力2列目(せん断)を1に上書きする。
    #       梁版(ケース列=2列目)からの流用と思われるが逐語再現する (壁は未検証)
    ch_wall_stress[:, 1] = 1

    # 水平応力部分
    H_stress = []
    H_ID = []
    for i_h in range(len(H_Lcase)):
        H_index = int(find_index(load_case_no_wall,
                                 load_case_no[int(H_Lcase[i_h]) - 1])) + 1
        if H_index == load_case_no_wall.size:
            H_stress.append(wall_stress[(load_case_index_wall[H_index - 1] + 1) * 2 - 1 - 1:, :])
            H_ID.append(wall_stress_ID[load_case_index_wall[H_index - 1]:, :])
        else:
            H_stress.append(wall_stress[(load_case_index_wall[H_index - 1] + 1) * 2 - 1 - 1:
                                        (load_case_index_wall[H_index] + 1) * 2 - 2, :])
            H_ID.append(wall_stress_ID[load_case_index_wall[H_index - 1]:
                                       load_case_index_wall[H_index], :])

    # せん断力を割増して合成する
    for i_h in range(len(H_Lcase)):
        for i_pm in range(1, pm_direction + 1):
            sub = H_stress[i_h].copy()
            sub[:, 1:3] = sub[:, 1:3] * qup_wall
            if i_pm == 1:
                sub[:, 0:6] = L_stress[:, 0:6] + sub[:, 0:6]
            else:
                sub[:, 0:6] = L_stress[:, 0:6] - sub[:, 0:6]
            ch_wall_stress = np.vstack([ch_wall_stress, sub])

    # 壁要素ID部分も合成する
    ch_wall_stress_ID = L_ID.copy()
    for i_h in range(len(H_Lcase)):
        for i_pm in range(1, pm_direction + 1):
            sub_ID = H_ID[i_h].copy()
            sub_ID[:, 2] = L_ID[0, 2] * 10 * i_pm + sub_ID[0, 2]
            ch_wall_stress_ID = np.vstack([ch_wall_stress_ID, sub_ID])
    ch_load_case_no_wall = _doublecheck_nonsort(ch_wall_stress_ID[:, 2])
    ch_load_case_index_wall = np.atleast_1d(find_index(
        ch_wall_stress_ID[:, 2], ch_load_case_no_wall))
    return (ch_wall_stress, ch_load_case_no_wall, ch_load_case_index_wall,
            ch_wall_stress_ID)


def _combine_stress_p(pm_direction, load_case_no, plate_stress,
                      load_case_no_plate, load_case_index_plate,
                      L_Lcase, H_Lcase, plate_up, pl_long):
    """analysis_case.m サブ関数 combine_stress_p の逐語移植.

    板要素応力は [要素番号, ケース, 節点番号, Fxy] の4列。
    長期部分は ×pl_long(0/1)、水平部分は ×plate_up して±合成。
    """
    plate_stress = np.atleast_2d(np.asarray(plate_stress, dtype=float))
    load_case_no_plate = np.asarray(load_case_no_plate, dtype=float).ravel()
    load_case_index_plate = np.asarray(load_case_index_plate).ravel()

    L_index = int(find_index(load_case_no_plate,
                             load_case_no[int(L_Lcase[0]) - 1])) + 1
    if L_index == load_case_no_plate.size:
        L_stress = plate_stress[load_case_index_plate[L_index - 1]:, :].copy()
    else:
        L_stress = plate_stress[load_case_index_plate[L_index - 1]:
                                load_case_index_plate[L_index], :].copy()
    L_stress[:, 3] = L_stress[:, 3] * pl_long
    ch_plate_stress = L_stress.copy()

    H_stress = []
    H_index_list = []
    for i_h in range(len(H_Lcase)):
        H_index = int(find_index(load_case_no_plate,
                                 load_case_no[int(H_Lcase[i_h]) - 1])) + 1
        H_index_list.append(H_index)
        if H_index == load_case_no_plate.size:
            H_stress.append(plate_stress[load_case_index_plate[H_index - 1]:, :])
        else:
            H_stress.append(plate_stress[load_case_index_plate[H_index - 1]:
                                         load_case_index_plate[H_index], :])

    for i_h in range(len(H_Lcase)):
        for i_pm in range(1, pm_direction + 1):
            sub = H_stress[i_h].copy()
            sub[:, 1] = (plate_stress[load_case_index_plate[L_index - 1], 1]
                         * 10 * i_pm
                         + plate_stress[load_case_index_plate[H_index_list[i_h] - 1], 1])
            sub[:, 3] = sub[:, 3] * plate_up
            if i_pm == 1:
                sub[:, 3] = L_stress[:, 3] + sub[:, 3]
            else:
                sub[:, 3] = L_stress[:, 3] - sub[:, 3]
            ch_plate_stress = np.vstack([ch_plate_stress, sub])

    ch_load_case_no_plate = _doublecheck_nonsort(ch_plate_stress[:, 1])
    ch_load_case_index_plate = np.atleast_1d(find_index(
        ch_plate_stress[:, 1], ch_load_case_no_plate))
    return ch_plate_stress, ch_load_case_no_plate, ch_load_case_index_plate


# ===========================================================================
# S_truss_ratio_analysis.m
# ===========================================================================

def S_truss_ratio_analysis(sectionsize, buck_length, stress, timecase, SN,
                           ele_no, maxratios, maxratios_text, section_no,
                           Lminus, Truss_times, LOAS_CASE_NAME,
                           pick_section_name):
    """鉄骨引張材（トラス要素）の断面検定ディスパッチャ (S_truss_ratio_analysis.m).

    戻り値: (ratio_output, maxratios, maxratios_text)
    """
    if _empty(Lminus) or find_index(np.asarray(Lminus)[:, 0], section_no) == -1:
        Lminus = 0
    else:
        Lm = np.asarray(Lminus, dtype=float)
        Lminus = Lm[find_index(Lm[:, 0], section_no), 1]
    Tt = np.asarray(Truss_times, dtype=float)
    if find_index(Tt[:, 0], section_no) == -1:
        Truss_times = 1.0
    else:
        Truss_times = Tt[find_index(Tt[:, 0], section_no), 1]

    sectionsize = np.asarray(sectionsize, dtype=float).ravel() * 1000
    maxratios = list(maxratios)
    n = max(sectionsize.shape)
    code = sectionsize[n - 2]
    stress = np.asarray(stress, dtype=float)

    if code == 1000:  # H鋼断面の検定
        ratio_output = TH_analysis(sectionsize[0:n - 2], buck_length,
                                   Truss_times * stress, timecase, SN)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TH_analysis_text(
                sectionsize[0:n - 2], buck_length, Truss_times * stress,
                timecase, SN, ele_no, section_no, LOAS_CASE_NAME,
                pick_section_name)
    elif code == 2000:  # 中実角断面の検定
        ratio_output = TSB_analysis(sectionsize[0:n - 2], buck_length,
                                    Truss_times * stress, timecase, SN, Lminus)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TSB_analysis_text(
                sectionsize[0:n - 2], Truss_times * stress, timecase, SN,
                ele_no, section_no, Lminus, LOAS_CASE_NAME, pick_section_name)
    elif code == 3000:  # 中実丸断面の検定
        ratio_output = TSR_analysis(sectionsize[0:n - 2], buck_length,
                                    Truss_times * stress, timecase, SN)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TSR_analysis_text(
                sectionsize[0:n - 2], Truss_times * stress, timecase, SN,
                ele_no, section_no, LOAS_CASE_NAME, pick_section_name)
    elif code == 4000:  # 鋼管断面の検定＜圧縮引張検定＞
        ratio_output = TP_analysis(sectionsize[0:n - 2], buck_length,
                                   Truss_times * stress, timecase, SN)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TP_analysis_text(
                sectionsize[0:n - 2], buck_length, Truss_times * stress,
                timecase, SN, ele_no, section_no, LOAS_CASE_NAME,
                pick_section_name)
    elif code == 5000:  # 溝形鋼断面の検定
        ratio_output = TC_analysis(sectionsize[0:n - 2], buck_length,
                                   Truss_times * stress, timecase, SN, Lminus)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TC_analysis_text(
                sectionsize[0:n - 2], Truss_times * stress, timecase, SN,
                ele_no, section_no, Lminus, LOAS_CASE_NAME, pick_section_name)
    elif code == 7000:  # 角形鋼管鋼断面の検定＜圧縮引張検定＞
        ratio_output = TBOX_analysis(sectionsize[0:n - 2], buck_length,
                                     Truss_times * stress, timecase, SN)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TBOX_analysis_text(
                sectionsize[0:n - 2], buck_length, Truss_times * stress,
                timecase, SN, ele_no, section_no, LOAS_CASE_NAME,
                pick_section_name)
    elif code == 9000:  # アングル断面の検定
        ratio_output = TL_analysis(sectionsize[0:n - 2], buck_length,
                                   Truss_times * stress, timecase, SN, Lminus)
        if np.max(np.max(ratio_output)) > max(np.atleast_1d(maxratios[1])):
            maxratios[0] = ele_no
            maxratios[1] = float(np.max(np.max(ratio_output)))
            maxratios_text = TL_analysis_text(
                sectionsize[0:n - 2], Truss_times * stress, timecase, SN,
                ele_no, section_no, Lminus, LOAS_CASE_NAME, pick_section_name)
    # 該当コードなし → MATLAB同様 ratio_output 未定義エラー
    return ratio_output, maxratios, maxratios_text


# ===========================================================================
# 結果コンテナ
# ===========================================================================

class CheckResult(object):
    """run_steel_check の結果.

    属性:
      beam_ratio      ndarray (3n×6): [要素番号, 検定ケース番号, cb, sz, sy, combi]
                      (要素ごと i端/中央/j端 の3行, MATLAB beam_ratio と同一構成)
      truss_ratio     ndarray (m×4): [要素番号, 検定ケース番号, 検定比(引張/圧縮)]
      maxratios       list[ndarray(断面数×3)] ケースごと
                      [断面番号, 最大検定比の要素番号, 最大検定比]
      maxratios_text  list[list[list[str]]] [断面index][ケースindex] → 検定詳細文
      lb_table        list[[断面番号, 要素番号, lb, iy, H, B, tw, tf]]
      section_nos     ndarray 使用断面番号 (ダミー材除く, 昇順)
      section_names   list[str] section_nos に対応する断面名
      load_case_no    ndarray 検定ケース番号 (ch_load_case_no_beam)
      LCNAME          list[str] 検定ケース名
      check_case_num  int 総検討荷重ケース数
      skipped         list[dict] スコープ外要素の記録
                      {'element_no','material_no','material_type','section_no',
                       'load_case','kind'}
      section_t_tex   list[str] 幅厚比検討のTeX (check_t_ratio)
      load_direction  ndarray 検定ケースの荷重方向
    """

    def __init__(self):
        self.beam_ratio = np.zeros((0, 6))
        self.truss_ratio = np.zeros((0, 4))
        self.maxratios = []
        self.maxratios_text = []
        self.lb_table = []
        self.section_nos = np.zeros(0)
        self.section_names = []
        self.load_case_no = np.zeros(0)
        self.LCNAME = []
        self.check_case_num = 0
        self.skipped = []
        self.section_t_tex = []
        self.load_direction = np.zeros((0, 3))
        self.element_info = np.zeros((0, 6))  # mgtopen_element の要素行列


_MAT_TYPE_NAME = {0: 'UNKNOWN', 1: 'S(SN)', 2: 'S(冷間成形)', 3: 'RC', 4: 'PC',
                  5: 'SRC', 6: 'CFT', 7: 'RC杭', 8: 'RM', 10: 'W'}


# ===========================================================================
# メインパイプライン
# ===========================================================================

# SRC検定を実行したことの検知 (KCUA照合済(2026-07-09)。未照合経路がある旨をUIに注記)
SRC_CHECK_NOTE = {'hit': False}


def run_steel_check(mgt_path, beam_stress_path, truss_stress_path=None,
                    select_length=1,
                    c_up={'STKR': 1.4, 'BCR': 1.3, 'BCP': 1.2},
                    H_beam_RCsupport=2, judge_H=2, limit_sec_no=9000,
                    qup=None, pm_direction=None, tr_long=None,
                    select_unit=float('inf'),
                    truss_bolt_loss=None, truss_times=None,
                    case_types=None, wal_up=1.0, w_material_map=None,
                    w_panel_co=None, rc_params=None, src_cover=40.0,
                    pile_cover=100.0, pc_params=None,
                    wall_stress_path=None, plate_stress_path=None,
                    plate_up=1.0, pl_long=None):
    """MIDAS mgt + 応力txt から鉄骨・木造の断面検定を実行する.

    (関数名は従来のまま run_steel_check だが、対象は鉄骨に加えて
     木造 material 2列目=10 の梁要素 (W_ratio_analysis: 中実角/中実丸) と
     トラス要素 (W_truss_ratio_analysis: Xブレース置換の壁面) に広がった)

    引数:
      mgt_path            MIDAS/Gen mgtファイル
      beam_stress_path    梁要素応力 (1:要素 2:ケース 3:Fx 4:Fy 5:Fz 6:Mx 7:My 8:Mz)
                          2行/要素形式の場合は3行(i/中央/j)へ展開する
      truss_stress_path   トラス要素応力 (1:要素 2:ケース 3:Fx(i) 4:Fx(j))
      select_length       1:実長指定 2:部材倍数指定 3:部材長 (getbuck)
      c_up                冷間成形角形鋼管の応力割増
                          dict {'STKR','BCR','BCP','BSH'(省略時1.0)} または配列
      H_beam_RCsupport    H形梁のRCスラブ拘束 (2:考慮しない)
      judge_H             H_analysis の judge_H
      limit_sec_no        この番号以上の断面はダミー材として検定しない
      qup / pm_direction / tr_long   qup_case / analysis_case の非対話設定
      select_unit         inf:結合部材を使わない (Midasratio GUI 相当), 0:一括表示
      truss_bolt_loss     dict {断面番号: 欠損幅mm} (既定: 欠損なし=0)
      truss_times         dict {断面番号: 応力倍率} (既定: 1.00)
      case_types          荷重ケース種別の明示指定 (case_types_to_lcase 参照)
                          {ケース番号: ('L'|'H'|'S', ケース名)}。
                          None (既定) はケース数からの自動判定
                          (l_case_ratio_analysis) — 従来互換。
                          引張専用材等の収束計算で長期+水平が組合せ済みの
                          応力ケースには 'S' を指定する (長期二重加算防止)。
      wal_up              木材の長期許容応力度の割増 (MIDASratioplot_fig.m
                          冒頭 questdlg「長期荷重の木部材の許容応力度は？」)。
                          1.0=長期 (既定) / 1.3=中長期 (積雪)。
      w_material_map      木材料の W_AL 行番号 (1-based) の明示指定
                          dict {mgt材料番号: W_AL行番号}。
                          未指定の木材料は材料名から自動対応 (w_al_match:
                          E***F***パターン・樹種名+等級・構造用合板)。
                          自動対応できない木材料が検定対象になった場合は
                          エラーとする (UIの木材種別選択で指定する)。
      w_panel_co          木造トラス要素 (筋かい・壁面のXブレース置換) の
                          壁倍率 dict {断面番号: 壁倍率}
                          (MATLAB版 w_panel_co.m の listdlg の引数化)。
      rc_params           RC検定の設定 (MIDASratioplot_fig.m のRC入力
                          ダイアログ群の引数化)。dict:
                            'rcw': {断面番号: {'mode':1(径とピッチ指定)|
                                    2(端部補強筋指定), 'sd':2(ダブル)|1(シングル),
                                    'v_di','v_pitch','h_di','h_pitch',
                                    'num_rebar'(mode2の片側本数)}}
                                    → RCW_input.m 相当で RCWcolumns を組立て
                            'rcg': {断面番号: 17列配筋行} 梁として配筋指定
                                    (RCG_input.m 相当。i端/中央/j端同一配筋で
                                     RCbeams へ追加合流)
                            'rcc': {断面番号: {'m_di','m_num','h_di','h_pitch',
                                    'h_num'}} 柱として配筋指定 (RCC_input.m
                                    相当で RCcolumns へ追加合流。RC柱検定は
                                    未移植のため検定到達時は明示エラー)
                            'wall_cover'/'beam_cover'/'column_cover': かぶり
                                    (既定40 = inputdlgの既定値)
                            'method_rcw': 1端部のみ/2縦筋のみ/3端部+縦筋
                                    (既定3 = questdlg既定「端部補強筋+縦筋」。
                                     壁式 walldesign_index=2 の面内曲げのみ使用)
                            'RCQ': 1応力割増/2部材端終局M/3小さい方 (既定1 =
                                    GUI既定 Value=2-1)
                            'walldesign_index': 1耐力壁付ラーメン/2壁式 (既定1。
                                    RC_ratio_analysis内で断面番号>2999は2,
                                    <2999は1に上書きされる点に注意)
                            'L_43': 梁引張鉄筋比割増(4/3)対象の断面番号list
                            'w_l': 壁有効長さ [断面番号,i端,中央?,…]x5列 または None
                            'wall_r': 壁せん断低減 [断面番号,低減率] または None
                          RCのせん断割増は qup 引数 (qup_case) で指定する
                          (MATLAB ルート1: {'beam':1.5,'wall':2.0})。

    戻り値: CheckResult
    """
    # ---- c_up を MATLAB互換の配列 [STKR BCR BCP BSH] へ ----
    # ユーザー承認による原典(MATLAB)バグ修正 2026-07-08:
    # 引張材(丸鋼)が圧縮のとき検定値が負となる原典バグの修正適用を検知する
    TSR_SIGN_FIX['hit'] = False
    SRC_CHECK_NOTE['hit'] = False
    RC4_YOSE_FIX['hit'] = False
    RCSR_NMZ_FIX['hit'] = False
    UNVERIFIED_NOTE['hit'] = set()

    if isinstance(c_up, dict):
        c_up_arr = np.array([c_up.get('STKR', 1.0), c_up.get('BCR', 1.0),
                             c_up.get('BCP', 1.0), c_up.get('BSH', 1.0)],
                            dtype=float)
    else:
        c_up_arr = np.asarray(c_up, dtype=float).ravel()

    # ---- モデルデータ読み込み ----
    node = mgtopen_node(mgt_path)
    element = mgtopen_element(mgt_path)
    sections, A_section_no, section_name = mgtopen_section(mgt_path)
    A_section_no = np.asarray(A_section_no, dtype=float).ravel()
    material = mgtopen_material(mgt_path, w_select=w_al_match)
    # 木材料の W_AL 行番号の明示指定 (材料名からの自動対応の上書き)
    if w_material_map:
        for m_no, w_idx in dict(w_material_map).items():
            m_idx = find_index(material[:, 0], float(m_no))
            if m_idx == -1:
                print('注意: w_material_map の材料番号 %s は mgt に存在しない'
                      'ため無視します' % m_no)
            elif material[m_idx, 1] != 10:
                print('注意: w_material_map の材料番号 %s は木材料ではない'
                      'ため無視します' % m_no)
            else:
                material[m_idx, 2] = float(w_idx)
    lky, lkz, lb = mgtopen_length(mgt_path)
    # ---- 板・壁要素まわりのモデル情報 (板要素・壁要素検定用) ----
    thickness = mgtopen_thickness(mgt_path)
    plate = mgtopen_plate(mgt_path)
    wall_element, wall_base = mgtopen_wall(mgt_path, node)
    if np.size(wall_element) > 0:
        RCwalls, _FL_name_w = mgtopen_RCwall(mgt_path, wall_element)
    else:
        RCwalls = np.array([])
    section_t_tex = check_t_ratio(sections, A_section_no, section_name)

    # ---- RC配筋情報の設定 (MIDASratioplot_fig.m RC入力ダイアログ群の引数化) ----
    rc_params = dict(rc_params or {})
    RCbeams = mgtopen_RCbeam(mgt_path)
    # UIで「梁」として配筋指定した断面を追加 (RCG_input.m 相当。
    # i端・中央・j端とも同一配筋の3行を登録する)
    for sec_no_key in sorted(rc_params.get('rcg') or {}, key=float):
        row17 = np.asarray(rc_params['rcg'][sec_no_key],
                           dtype=float).ravel()
        if row17.size != 17:
            raise ValueError('RC梁配筋指定(断面%s)の列数が不正です: %d'
                             % (sec_no_key, row17.size))
        RCbeams.append([float(sec_no_key), np.tile(row17, (3, 1))])
    if RCbeams:
        RCbeam_secNO = np.asarray([row[0] for row in RCbeams], dtype=float)
    else:
        RCbeam_secNO = np.zeros((0,))
    RCcolumns = mgtopen_RCcolumn(mgt_path)
    # UIで「柱」として配筋指定した断面を追加 (RCC_input.m 92行相当:
    # [断面番号,1,総本数,主筋径,本数/4+1,本数/4+1,HOOP径,ピッチ,本数y,本数z])
    _rcc = rc_params.get('rcc') or {}
    if _rcc:
        _rcc_rows = []
        for sec_no_key in sorted(_rcc, key=float):
            c = _rcc[sec_no_key]
            _rcc_rows.append([
                float(sec_no_key), 1.0, float(c['m_num']), float(c['m_di']),
                float(c['m_num']) / 4 + 1, float(c['m_num']) / 4 + 1,
                float(c['h_di']), float(c['h_pitch']),
                float(c['h_num']), float(c['h_num'])])
        _base = np.atleast_2d(np.asarray(RCcolumns, dtype=float))
        if _base.size:
            RCcolumns = np.vstack([_base, np.asarray(_rcc_rows,
                                                     dtype=float)])
        else:
            RCcolumns = np.asarray(_rcc_rows, dtype=float)
    rcw_settings = rc_params.get('rcw') or {}
    RCWcolumns = []
    _sb_table = sections[2] if (len(sections) > 2 and sections[2] is not None
                                and np.asarray(sections[2]).size) else np.zeros((0, 3))
    for sec_no_key in sorted(rcw_settings, key=float):
        s = rcw_settings[sec_no_key]
        RCWcolumns.append(rcw_input_row(
            float(sec_no_key), _sb_table,
            int(s.get('mode', 1)), int(s.get('sd', 2)),
            float(s.get('v_di', 10)), float(s.get('v_pitch', 200)),
            float(s.get('h_di', 10)), float(s.get('h_pitch', 200)),
            num_rebar=float(s.get('num_rebar', 2))))
    if RCWcolumns:
        RCWcolumns = np.asarray(RCWcolumns, dtype=float)
    else:
        RCWcolumns = np.zeros((0, 11))
    beam_cover = float(rc_params.get('beam_cover', 40))
    wall_cover = float(rc_params.get('wall_cover', 40))
    column_cover = float(rc_params.get('column_cover', 40))
    method_rcw = int(rc_params.get('method_rcw', 3))  # questdlg既定「端部補強筋+縦筋」
    RCQ = int(rc_params.get('RCQ', 1))  # Midasratio GUI既定 Value=2 → RCQ=1
    walldesign_index_rc = int(rc_params.get('walldesign_index', 1))
    _L43_list = rc_params.get('L_43') or []
    if len(_L43_list):
        L_43 = np.column_stack([np.asarray(_L43_list, dtype=float),
                                np.ones(len(_L43_list))])
    else:
        L_43 = None
    w_l = rc_params.get('w_l')
    wall_r = rc_params.get('wall_r')

    # ---- SRC配筋情報の設定 (MIDASratioplot_fig.m 71-104行) ----
    #  SRC配筋は mgt (*SRC-BEAM/*SRC-COLUMN) から読む。ユーザー入力は
    #  コンクリートかぶり厚 src_cover (原典 inputdlg 既定40) のみ。
    SRCbeams = mgtopen_SRCbeam(mgt_path)
    if SRCbeams:
        SRCbeam_secNO = np.asarray([b[0] for b in SRCbeams], dtype=float)
    else:
        SRCbeam_secNO = np.zeros((0,))
    SRCcolumns = mgtopen_SRCcolumn(mgt_path)
    cover_src = float(src_cover)
    if np.asarray(SRCcolumns).size > 0:
        SRC_ALW_NM = SRCcolumn_ALNM(SRCcolumns, sections, material,
                                    element, cover_src)
    else:
        SRC_ALW_NM = []
    if SRCbeams:
        SRC_ALW_M = SRCbeam_ALM(SRCbeams, sections, material,
                                element, cover_src)
    else:
        SRC_ALW_M = []

    # ---- 梁要素の応力読み込み ----
    beam_stress = loadtxt_tolerant(beam_stress_path)
    beam_stress = np.atleast_2d(beam_stress)

    # 2行/要素形式（終局）なら3行へ展開 (MIDASratioplot_fig.m)
    if beam_stress[0, 0] == beam_stress[2, 0]:
        pass
    else:
        num_b = beam_stress.shape[0] // 2
        new_bs = np.zeros((3 * num_b, beam_stress.shape[1]))
        for i in range(1, num_b + 1):
            new_bs[3 * num_b - 3 * i + 0, :] = beam_stress[2 * num_b - 2 * i + 0, :]
            new_bs[3 * num_b - 3 * i + 1, :] = (
                beam_stress[2 * num_b - 2 * i + 0, :] * 0.5
                + beam_stress[2 * num_b - 2 * i + 1, :] * 0.5)
            new_bs[3 * num_b - 3 * i + 2, :] = beam_stress[2 * num_b - 2 * i + 1, :]
        beam_stress = new_bs

    # ---- トラス要素の応力読み込み ----
    if truss_stress_path is None:
        truss_stress = None
    else:
        truss_stress = np.atleast_2d(loadtxt_tolerant(truss_stress_path))
        if truss_stress.size == 0:
            truss_stress = None

    # ---- 荷重ケースが昇順ブロックで並んでいない応力ファイルへの対応 ----
    # (後段の find_index によるブロック切り出しはケース昇順の並びを前提と
    #  するため、降順箇所があればケース番号で安定ソートする。行の並びが
    #  既に昇順のファイルでは何もしない)
    if beam_stress.shape[0] > 1 and np.any(np.diff(beam_stress[:, 1]) < 0):
        beam_stress = beam_stress[np.argsort(beam_stress[:, 1],
                                             kind='stable')]
    if (truss_stress is not None and truss_stress.shape[0] > 1
            and np.any(np.diff(truss_stress[:, 1]) < 0)):
        truss_stress = truss_stress[np.argsort(truss_stress[:, 1],
                                               kind='stable')]

    # ---- 応力データの荷重ケース数と区切り番号読み込み ----
    load_case_no_beam = doublecheck(beam_stress[:, 1])
    load_case_index_beam = find_index(beam_stress[:, 1], load_case_no_beam)
    if truss_stress is None:
        load_case_index_truss = None
        load_case_no_truss = None
    else:
        load_case_no_truss = doublecheck(truss_stress[:, 1])
        load_case_index_truss = find_index(truss_stress[:, 1], load_case_no_truss)

    # ---- PC導入応力の加算 (MIDASratioplot_fig.m 550-556) ----
    #  プレストレス不静定応力(pc_stress)×有効率を長期(先頭ケース)へ加算する
    if pc_params and not pc_params.get('skip'):
        _pcs = pc_params.get('pc_stress')
        if _pcs is not None and np.size(_pcs):
            _pcs = np.atleast_2d(np.asarray(_pcs, dtype=float))
            _end1 = (int(load_case_index_beam[1])
                     if np.size(load_case_index_beam) > 1
                     else beam_stress.shape[0])
            if _pcs.shape[0] != _end1:
                raise ValueError(
                    'pc_stressの行数(%d)が長期ケースのbeam_stress行数(%d)と'
                    '一致しません' % (_pcs.shape[0], _end1))
            beam_stress[0:_end1, 2:8] = (beam_stress[0:_end1, 2:8]
                                         + float(pc_params.get('PC_eff', 0.85))
                                         * _pcs[:, 2:8])

    # ---- 板要素応力の読込 (MIDASratioplot_fig.m 482-508) ----
    plate_stress = None
    load_case_no_plate = None
    load_case_index_plate = None
    if plate_stress_path and np.size(plate) > 0:
        _ps = np.atleast_2d(loadtxt_tolerant(plate_stress_path))
        if _ps.size:
            if _ps.shape[1] < 4:
                raise ValueError('plate_stressファイルの列数が不足しています '
                                 '(要素番号, 荷重ケース, 節点番号, Fxy の4列)')
            plate_stress = _ps
            load_case_no_plate = doublecheck(plate_stress[:, 1])
            load_case_index_plate = find_index(plate_stress[:, 1],
                                               load_case_no_plate)
    elif plate_stress_path:
        print('注意: mgtに板要素(*ELEMENT PLATE)が無いため plate_stress は'
              '無視します')

    # ---- 壁要素応力の読込 (MIDASratioplot_fig.m 511-536) ----
    wall_stress = None
    wall_stress_ID = None
    load_case_no_wall = None
    load_case_index_wall = None
    if wall_stress_path and np.size(wall_element) > 0:
        from .draw_stress import get_wstress
        raw = get_wstress(wall_stress_path)
        if np.size(raw):
            n_pair = raw.shape[0] // 2
            wall_stress_ID = np.zeros((n_pair, 3))
            _ws = np.zeros((2 * n_pair, 6))
            for _i in range(n_pair):
                wall_stress_ID[_i, :] = raw[2 * _i, 0:3]
                _ws[2 * _i, :] = raw[2 * _i, 3:9]
                _ws[2 * _i + 1, :] = raw[2 * _i + 1, 0:6]
            wall_stress = _ws
            load_case_no_wall = doublecheck(wall_stress_ID[:, 2])
            load_case_index_wall = find_index(wall_stress_ID[:, 2],
                                              load_case_no_wall)
    elif wall_stress_path:
        print('注意: mgtに壁要素(*ELEMENT WALL)が無いため wall_stress は'
              '無視します')

    num_load_case = max(
        load_case_no_beam.shape[0],
        0 if load_case_no_truss is None else load_case_no_truss.shape[0],
        0 if load_case_no_plate is None else load_case_no_plate.shape[0],
        0 if load_case_no_wall is None else load_case_no_wall.shape[0])
    _lc_all = [load_case_no_beam]
    if load_case_no_truss is not None:
        _lc_all.append(load_case_no_truss)
    if load_case_no_plate is not None:
        _lc_all.append(load_case_no_plate)
    if load_case_no_wall is not None:
        _lc_all.append(load_case_no_wall)
    load_case_no = doublecheck(np.concatenate(_lc_all))

    # ---- 荷重ケース種別判定・割増オプション・検定ケース作成 ----
    if case_types is None:  # 自動判定 (従来互換)
        (L_Lcase, H_Lcase, S_Lcase, load_case_name,
         load_direction) = l_case_ratio_analysis(num_load_case, load_case_no)
    else:  # ケース種別の明示指定
        (L_Lcase, H_Lcase, S_Lcase, load_case_name,
         load_direction) = case_types_to_lcase(num_load_case, load_case_no,
                                               case_types)
    qup_beam, qup_wall, qup_beam_src, pm_direction = qup_case(
        L_Lcase, H_Lcase, S_Lcase, qup=qup, pm_direction=pm_direction)
    _raw_case_no = np.asarray(load_case_no, dtype=float).copy()  # 合成前

    (load_case_no, ch_beam_stress, ch_load_case_no_beam,
     ch_load_case_index_beam, ch_truss_stress, ch_load_case_no_truss,
     ch_load_case_index_truss, LCNAME, check_case_num,
     load_direction,
     ch_wall_stress, ch_load_case_no_wall, ch_load_case_index_wall,
     ch_wall_stress_ID, ch_plate_stress, ch_load_case_no_plate,
     ch_load_case_index_plate) = analysis_case(
        L_Lcase, H_Lcase, S_Lcase, qup_beam, qup_wall, pm_direction,
        load_case_no, beam_stress, load_case_no_beam, load_case_index_beam,
        truss_stress, load_case_no_truss, load_case_index_truss,
        load_case_name, num_load_case, load_direction, tr_long=tr_long,
        wall_stress=wall_stress, load_case_no_wall=load_case_no_wall,
        load_case_index_wall=load_case_index_wall,
        wall_stress_ID=wall_stress_ID,
        plate_stress=plate_stress, load_case_no_plate=load_case_no_plate,
        load_case_index_plate=load_case_index_plate,
        plate_up=plate_up, pl_long=pl_long)

    # MIDASratioplot_fig.m: load_direction = [doublecheck(ch(:,2)) load_direction]
    ld_cases = doublecheck(ch_beam_stress[:, 1]).reshape(-1, 1)
    if np.asarray(load_direction).size == ld_cases.shape[0] * 2:
        load_direction = np.hstack([ld_cases, np.asarray(load_direction,
                                                         dtype=float)])
    else:
        load_direction = ld_cases

    # ---- 断面別最大検定値保存用セルの設定 ----
    section_nos = doublecheck(element[:, 2])
    section_nos = section_nos[section_nos < limit_sec_no]
    num_section = section_nos.shape[0]

    # モデルにおいて利用されている厚さ種類数の取得 (壁→板の順に結合)
    if ch_wall_stress is None:
        num_thick = 0
        num_thick_wall = 0
        thick_nos = np.zeros(0)
        thickness_wall = np.zeros((0, 3))
    else:
        thick_nos = doublecheck(wall_element[:, 7])
        thickness_wall = np.atleast_2d(thickness)[
            np.atleast_1d(find_index(thickness[:, 0], thick_nos)), :]
        num_thick = thick_nos.shape[0]
    num_thick_wall = num_thick
    if ch_plate_stress is not None:
        thick_nos_pl = doublecheck(plate[:, 2])
        thick_nos = np.concatenate([thick_nos, thick_nos_pl])
        num_thick = thick_nos.shape[0]

    maxratios = []
    for i in range(check_case_num):
        m = np.zeros((num_section + num_thick, 3))
        m[:, 0] = np.concatenate([section_nos, thick_nos])
        maxratios.append(m)
    maxratios_text = [[[] for _ in range(check_case_num)]
                      for _ in range(num_section + num_thick)]

    # ---- 分割部材の一括表示オプション ----
    unit_element = mgtopen_unit_2015(mgt_path, element, node)
    for iunit in range(len(unit_element)):
        unit_ele_index = np.atleast_1d(
            find_index(element[:, 0], unit_element[iunit][1]))
        unit_ele_sec_no = element[unit_ele_index, 2]
        unit_ele_sec_no = (unit_ele_sec_no < limit_sec_no).astype(float)
        s = unit_ele_sec_no.sum()
        p = unit_ele_sec_no.prod()
        if s == 0 and p == 0:
            pass
        elif s > 0 and p > 0:
            pass
        elif s != p:
            Lu = unit_ele_sec_no.shape[0]
            if unit_ele_sec_no[1:Lu - 1].prod() == 0:
                raise RuntimeError('unit_element: 中間にダミー材')  # MATLAB: stop
            else:
                nodes_u = list(np.atleast_1d(unit_element[iunit][2]))
                eles_u = np.atleast_1d(unit_element[iunit][1])
                if unit_ele_sec_no[Lu - 1] == 0:
                    del nodes_u[Lu]  # 端部の節点番号削除
                if unit_ele_sec_no[0] == 0:
                    del nodes_u[0]
                unit_element[iunit][2] = np.asarray(nodes_u)
                unit_element[iunit][1] = eles_u[unit_ele_sec_no != 0]

    uniele_ID = []
    for i_u in range(len(unit_element)):
        uniele_ID.extend(np.atleast_1d(unit_element[i_u][1]).ravel().tolist())
    uniele_ID = sorted(uniele_ID)
    uniele_ID = np.asarray(uniele_ID, dtype=float).reshape(-1, 1)

    skipped = []
    missing_ele_warned = set()  # mgtに存在しない応力ファイル要素の警告済み番号

    # ---- RC断面の配筋定義漏れチェック (mgt/設定側の問題を先に指摘) ----
    rc_missing = []
    for e_no in doublecheck(ch_beam_stress[:, 0]):
        eidx = find_index(element[:, 0], e_no)
        if eidx == -1:
            continue
        m_idx = find_index(material[:, 0], element[eidx, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        s_no = element[eidx, 2]
        if s_no >= limit_sec_no:
            continue
        in_beam = RCbeam_secNO.size > 0 and find_index(RCbeam_secNO, s_no) != -1
        in_wall = RCWcolumns.size > 0 and find_index(RCWcolumns[:, 0], s_no) != -1
        in_col = (RCcolumns is not None and np.asarray(RCcolumns).size > 0
                  and find_index(np.atleast_2d(np.asarray(RCcolumns, dtype=float))[:, 0], s_no) != -1)
        if not (in_beam or in_wall or in_col) and int(s_no) not in rc_missing:
            rc_missing.append(int(s_no))
    if rc_missing:
        names = []
        for s_no in sorted(rc_missing):
            si = find_index(A_section_no, s_no)
            names.append('%d(%s)' % (s_no, section_name[si] if si != -1 else '?'))
        raise ValueError(
            'RC断面の配筋情報が未定義です: ' + '、'.join(names)
            + '。検定タブの「RC配筋設定」で種別(壁/梁/柱)と配筋を指定するか、'
              'MIDASの配筋(*REBAR-BEAM/*REBAR-COLUMN)を設定してください。')

    # =======================================================================
    # 梁要素の断面検定
    # =======================================================================
    num_beam = ch_beam_stress.shape[0]
    beam_ratio = np.zeros((num_beam, 6))
    lb_table = []

    cb_judge = np.array([0.0, 0.0])  # トラスループでのMATLAB残置変数の初期値
    for i in range(1, num_beam // 3 + 1):
        r0 = 3 * i - 3
        beam_ratio[r0:r0 + 3, 0:2] = ch_beam_stress[r0:r0 + 3, 0:2]
        element_no = ch_beam_stress[r0, 0]

        # 応力ファイルにあるがmgtに無い要素は検定不能のためスキップ (検定比0)
        if find_index(element[:, 0], element_no) == -1:
            if int(element_no) not in missing_ele_warned:
                missing_ele_warned.add(int(element_no))
                print('mgt記述の注意: 応力ファイルの要素%dはmgtファイルに'
                      '存在しないため検定をスキップしました' % int(element_no))
            continue

        # 断面情報取得
        sectionsize = sectionget(element_no, element, sections, limit_sec_no)

        if sectionsize[0] < limit_sec_no:
            load_case = ch_beam_stress[r0, 1]
            load_no = find_index(ch_load_case_no_beam, load_case)
            eidx = find_index(element[:, 0], element_no)
            cb_judge = np.array([
                column_beam_judge([eidx], element, node)[0], element[eidx, 5]])

            section_no = sectionsize[0]
            pick_section_name = pick_text(
                section_name[find_index(A_section_no, section_no)], '_', -1)
            sectionsize = sectionsize[1:]
            section_index = find_index(section_nos, section_no)

            # 材料情報取得
            material_no = materialget(element_no, element)
            material_index = find_index(material[:, 0], material_no)
            ele_material = material[material_index, :]

            # 応力と長短期設定 (3:Nx 4:Fy 5:Fz 7:My 8:Mz)
            stress = ch_beam_stress[r0:r0 + 3][:, [2, 3, 4, 6, 7]]
            timecase = ch_beam_stress[r0, 1]
            # MATLABのLOAD_CASE_NAMEはstrvcatの固定幅char行のため
            # 全材料共通で最大ケース名幅まで空白パディングする
            LOAS_CASE_NAME = LCNAME[find_index(ch_load_case_no_beam, timecase)]
            LOAS_CASE_NAME = LOAS_CASE_NAME.ljust(max(len(x) for x in LCNAME))

            # 部材検定用長さの設定
            ij_select = 0
            ij_reverse = 0
            if _empty(uniele_ID):
                nodei_index, nodej_index = node_index_get(element_no, element, node)
                ele_length = float(np.linalg.norm(
                    node[nodei_index, 1:4] - node[nodej_index, 1:4]))
                buck_length = getbuck(ele_length, element_no, lky, lkz, lb,
                                      select_length, cb_judge,
                                      H_beam_RCsupport, stress)
            else:
                judge_unit = find_index(uniele_ID[:, 0], element_no) + 1  # MATLAB値
                if judge_unit <= select_unit:
                    nodei_index, nodej_index = node_index_get(element_no,
                                                              element, node)
                    ele_length = float(np.linalg.norm(
                        node[nodei_index, 1:4] - node[nodej_index, 1:4]))
                    buck_length = getbuck(ele_length, element_no, lky, lkz, lb,
                                          select_length, cb_judge,
                                          H_beam_RCsupport, stress)
                else:  # 結合部材の長さをベースに決定する
                    for iunit in range(len(unit_element)):
                        in_unit_element = np.atleast_1d(unit_element[iunit][1])
                        IDinunit = find_index(in_unit_element, element_no)
                        if IDinunit == -1:
                            continue
                        unit_nodes = np.atleast_1d(unit_element[iunit][2])
                        nodei_no = unit_nodes[0]
                        nodej_no = unit_nodes[-1]
                        nodei_index = find_index(node[:, 0], nodei_no)
                        nodej_index = find_index(node[:, 0], nodej_no)
                        ele_length = float(np.linalg.norm(
                            node[nodei_index, 1:4] - node[nodej_index, 1:4]))

                        # RC配筋のi端/j端選択 (MIDASratioplot_fig.m 1085-1103行)
                        ni2, nj2 = node_index_get(element_no, element, node)
                        if (find_index(unit_nodes, node[ni2, 0])
                                < find_index(unit_nodes, node[nj2, 0])):
                            ij_reverse = 1
                        else:
                            ij_reverse = -1
                        nue = in_unit_element.shape[0]
                        fiu = find_index(in_unit_element, element_no) + 1  # MATLAB 1-based
                        if fiu > nue - math.ceil(nue / 3) + 1:
                            ij_select = 3
                        elif fiu == nue - math.ceil(nue / 3) + 1:
                            ij_select = 2.5
                        elif fiu < math.ceil(nue / 3):
                            ij_select = 1
                        elif fiu == math.ceil(nue / 3):
                            ij_select = 1.5
                        else:
                            ij_select = 2

                        buck_length = getbuck_uni(
                            ele_length, element_no, lky, lkz, lb,
                            select_length, cb_judge, H_beam_RCsupport,
                            ch_beam_stress, timecase, unit_nodes,
                            in_unit_element)

            if ele_material[1] == 1 or ele_material[1] == 2:  # 鉄骨の検定
                # 長期荷重のpickup（冷間成形の応力UP用）
                if timecase == 1:
                    stress_L = stress
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    stress_L = ch_beam_stress[QL_ID - 1][:, [2, 3, 4, 6, 7]]

                (ratio_one, mr, mtext, lb_table) = S_ratio_analysis(
                    sectionsize, buck_length, stress, timecase,
                    ele_material[1:4], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, stress_L, c_up_arr, H_beam_RCsupport,
                    cb_judge, LOAS_CASE_NAME, lb_table, pick_section_name,
                    judge_H)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
            elif ele_material[1] == 3:  # RCの検定
                # MATLABのLOAS_CASE_NAMEはstrvcatで最大幅まで空白パディング
                # された char 行のため、RC検定詳細文の再現用に同じ形にする
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_rc = LOAS_CASE_NAME.ljust(_lc_w)
                # 長期荷重のpickup（応力割増用）
                if timecase == 1:
                    QL = stress[:, 0:5]
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_beam_stress[QL_ID - 1][:, [2, 3, 4, 6, 7]]

                # RCの検定(梁は曲げ・せん断の2列、壁柱はNMy/NMz/Qz/Qyの4列)
                (ratio_one, mr, mtext) = RC_ratio_analysis(
                    sectionsize, ele_length, stress, timecase,
                    ele_material[1:3], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, RCcolumns, RCbeams, RCWcolumns, QL,
                    qup_beam, qup_wall, column_cover, beam_cover,
                    wall_cover, RCbeam_secNO, wall_r, node, element,
                    load_direction, load_no, ij_select, ij_reverse,
                    LOAS_CASE_NAME_rc, w_l, L_43, RCQ, buck_length,
                    pick_section_name, walldesign_index_rc, method_rcw)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
            elif ele_material[1] == 10:
                # 木造の検定 (MIDASratioplot_fig.m: W_ratio_analysis)
                (ratio_one, mr, mtext) = W_ratio_analysis(
                    sectionsize, buck_length, stress, timecase,
                    ele_material[2], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, LOAS_CASE_NAME, pick_section_name, wal_up)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
            elif ele_material[1] == 5:  # SRCの検定
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_src = LOAS_CASE_NAME.ljust(_lc_w)
                # 長期荷重のpickup。SRCは QL=[Qy,Qz] の2列 (RCと異なる)
                if timecase == 1:
                    QL = stress[:, 1:3]
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_beam_stress[QL_ID - 1][:, [3, 4]]

                (ratio_one, mr, mtext) = SRC_ratio_analysis(
                    sectionsize, ele_length, stress, timecase,
                    ele_material[2:4], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, SRCcolumns, SRCbeams, SRCbeam_secNO,
                    SRC_ALW_NM, SRC_ALW_M, QL, qup_beam_src, cover_src,
                    LOAS_CASE_NAME_src, RCQ, pick_section_name)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
                SRC_CHECK_NOTE['hit'] = True
            elif ele_material[1] == 8:  # RM(補強組積造)の検定
                from .rm_check import RM_ratio_analysis
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_rm = LOAS_CASE_NAME.ljust(_lc_w)
                if timecase == 1:
                    QL = stress[:, 0:5]
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_beam_stress[QL_ID - 1][:, [2, 3, 4, 6, 7]]
                (ratio_one, mr, mtext) = RM_ratio_analysis(
                    sectionsize, ele_length, stress, timecase,
                    ele_material[1:3], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, RCcolumns, RCbeams, RCWcolumns, QL,
                    qup_beam, qup_wall, column_cover, beam_cover,
                    wall_cover, RCbeam_secNO, wall_r, node, element,
                    load_direction, load_no, ij_select, ij_reverse,
                    LOAS_CASE_NAME_rm, w_l, L_43, RCQ, buck_length,
                    pick_section_name, walldesign_index_rc)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('RM(補強組積造)')

            elif ele_material[1] == 7:  # RC杭の検定
                from .pile_check import RCpile_ratio_analysis
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_pl = LOAS_CASE_NAME.ljust(_lc_w)
                if timecase == 1:
                    QL = stress[:, 0:5]
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_beam_stress[QL_ID - 1][:, [2, 3, 4, 6, 7]]
                (ratio_one, mr, mtext) = RCpile_ratio_analysis(
                    sectionsize, ele_length, stress, timecase,
                    ele_material[1:3], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, RCcolumns, QL, pile_cover,
                    LOAS_CASE_NAME_pl, RCQ, pick_section_name)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('RC杭')

            elif ele_material[1] == 4:  # PCの検定
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_pc = LOAS_CASE_NAME.ljust(_lc_w)
                if timecase == 1:
                    QL = stress[:, 0:5]
                else:
                    base = num_beam // ch_load_case_no_beam.shape[0]
                    QL_ID = np.array([(r % base) for r in
                                      range(3 * i - 2, 3 * i + 1)])
                    for iq in range(3):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_beam_stress[QL_ID - 1][:, [2, 3, 4, 6, 7]]
                if pc_params is None:
                    raise ValueError(
                        'PC部材(材料タイプ4, 要素%d)の検定にはPC情報の入力が'
                        '必要です。検定タブの「PC検定の設定」で有効率・設計'
                        'タイプ・ケーブル情報等を入力するか、「PC部材の検定を'
                        '行わない」を選択してください。' % int(element_no))
                if pc_params.get('skip'):
                    # 原典 PC_switch=0: 検定値100のダミーを出力
                    ratio_one = np.zeros((3, 4)) + 100
                    maxratios[load_no][section_index, 1:3] = 100
                    maxratios_text[section_index][load_no] = []
                else:
                    from .pc_check import PC_ratio_analysis
                    (ratio_one, mr, mtext) = PC_ratio_analysis(
                        sectionsize, ele_length, stress, timecase,
                        ele_material[1:3], element_no,
                        maxratios[load_no][section_index, 1:3],
                        maxratios_text[section_index][load_no],
                        section_no, RCcolumns, RCbeams, QL, qup_beam,
                        column_cover, beam_cover, RCbeam_secNO,
                        node, element, load_direction, load_no,
                        ij_select, ij_reverse, LOAS_CASE_NAME_pc, L_43,
                        pc_params.get('PC_slab'),
                        float(pc_params.get('PC_eff', 0.85)),
                        pc_params.get('pc_stress'),
                        pc_params.get('PC_type', 1),
                        float(pc_params.get('DL_ratio', 0.6)),
                        pc_params.get('cable_delta'),
                        pc_params.get('cable_select'),
                        pick_section_name)
                    maxratios[load_no][section_index, 1:3] = mr
                    maxratios_text[section_index][load_no] = mtext
                    UNVERIFIED_NOTE['hit'].add('PC')

            elif ele_material[1] == 6:  # CFTの検定
                from .cft_check import CFT_ratio_analysis
                _lc_w = max(len(x) for x in LCNAME)
                LOAS_CASE_NAME_cf = LOAS_CASE_NAME.ljust(_lc_w)
                # NOTE: 原典はQL(2列)をpickupするがCFT_ratio_analysisは
                #       QLを受けないため計算に使われない (省略)
                (ratio_one, mr, mtext) = CFT_ratio_analysis(
                    sectionsize, buck_length, stress, timecase,
                    ele_material[2:4], element_no,
                    maxratios[load_no][section_index, 1:3],
                    maxratios_text[section_index][load_no],
                    section_no, LOAS_CASE_NAME_cf, pick_section_name)
                maxratios[load_no][section_index, 1:3] = mr
                maxratios_text[section_index][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('CFT')

            else:
                # 未対応材料 → スキップ記録
                skipped.append({
                    'element_no': int(element_no),
                    'material_no': int(material_no),
                    'material_type': int(ele_material[1]),
                    'material_type_name': _MAT_TYPE_NAME.get(
                        int(ele_material[1]), str(int(ele_material[1]))),
                    'section_no': int(section_no),
                    'load_case': int(load_case),
                    'kind': 'beam'})
                ratio_one = np.zeros((3, 4))

            beam_ratio[r0:r0 + 3, 2:6] = ratio_one
        else:
            beam_ratio[r0:r0 + 3, 2:6] = np.zeros((3, 4))

    # =======================================================================
    # 板要素の断面検定 (MIDASratioplot_fig.m 1276-1340)
    # =======================================================================
    plate_ratio = np.zeros((0, 4))
    if ch_plate_stress is not None and np.size(ch_plate_stress):
        from .plate_check import (S_plate_ratio_analysis,
                                  RC_plate_ratio_analysis,
                                  W_plate_ratio_analysis)
        thick_nos_arr = np.asarray(thick_nos, dtype=float).ravel()
        num_plate = ch_plate_stress.shape[0]
        plate_ratio = np.zeros((num_plate, 4))
        plate_ratio[:, 0:3] = ch_plate_stress[:, 0:3]
        for i in range(1, num_plate + 1):
            load_case = ch_plate_stress[i - 1, 1]
            load_no = int(find_index(ch_load_case_no_plate, load_case))
            plate_no = ch_plate_stress[i - 1, 0]

            # 断面情報取得
            pidx = find_index(plate[:, 0], plate_no)
            if pidx == -1:
                print('mgt記述の注意: plate_stressの板要素%dはmgtに存在'
                      'しないためスキップしました' % int(plate_no))
                continue
            thick_no = plate[pidx, 2]
            material_no = plate[pidx, 1]
            thick_index = int(find_index(thick_nos_arr, thick_no))
            # NOTE: 原典1302-1303行は thick_nos 上の位置で thickness 表を
            #       引く (表の並びと thick_nos の並びが異なると別の厚さを
            #       参照しうる原典の癖をそのまま再現)
            element_thick = np.array(
                np.atleast_2d(thickness)[thick_index, 0:2], copy=True)
            thick_no2 = element_thick[0]
            element_thick = float(element_thick[1])

            material_index = find_index(material[:, 0], material_no)
            ele_material = material[material_index, :]

            stress_p = ch_plate_stress[i - 1, 3]
            timecase = ch_plate_stress[i - 1, 1]
            _lc_w = max(len(x) for x in LCNAME)
            LOAS_CASE_NAME = LCNAME[int(find_index(
                ch_load_case_no_plate, timecase))].ljust(_lc_w)
            row = num_section + thick_index
            ratio_one = 0.0
            if ele_material[1] in (1, 2):  # 鉄骨板の検定
                (ratio_one, mr, mtext) = S_plate_ratio_analysis(
                    element_thick, stress_p, timecase, ele_material[1:4],
                    thick_no2, maxratios[load_no][row, 1:3],
                    maxratios_text[row][load_no], LOAS_CASE_NAME, plate_no)
                maxratios[load_no][row, 1:3] = mr
                maxratios_text[row][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('鉄骨板要素')
            elif ele_material[1] == 3:  # RC板の検定
                (ratio_one, mr, mtext) = RC_plate_ratio_analysis(
                    element_thick, stress_p, timecase, ele_material[2],
                    thick_no2, maxratios[load_no][row, 1:3],
                    maxratios_text[row][load_no], LOAS_CASE_NAME, plate_no)
                maxratios[load_no][row, 1:3] = mr
                maxratios_text[row][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('RC板要素')
            elif ele_material[1] == 10:  # 木造面材壁の検定
                (ratio_one, mr, mtext) = W_plate_ratio_analysis(
                    element_thick, stress_p, timecase, ele_material[2],
                    thick_no2, maxratios[load_no][row, 1:3],
                    maxratios_text[row][load_no], LOAS_CASE_NAME, plate_no)
                maxratios[load_no][row, 1:3] = mr
                maxratios_text[row][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('木造面材壁(板要素)')
            # NOTE: 原典はSRC(4)/CFT(5)は「検定(未設定)」で素通し
            plate_ratio[i - 1, 3] = ratio_one

    # =======================================================================
    # 壁要素の断面検定 (MIDASratioplot_fig.m 1342-1421)
    # =======================================================================
    wall_ratio = np.zeros((0, 7))
    if ch_wall_stress is not None and np.size(ch_wall_stress):
        from .wall_check import RCwall_ratio_analysis
        from .draw_stress import _wallinfo_get
        thick_nos_arr = np.asarray(thick_nos, dtype=float).ravel()
        num_wall = ch_wall_stress_ID.shape[0]
        wall_ratio = np.zeros((2 * num_wall, 7))
        for i in range(1, num_wall + 1):
            wall_ratio[2 * i - 2:2 * i, 0:3] = np.vstack(
                [ch_wall_stress_ID[i - 1, :], ch_wall_stress_ID[i - 1, :]])
            load_case = ch_wall_stress_ID[i - 1, 2]
            load_no = int(find_index(ch_load_case_no_wall, load_case))

            select_wall, _swi = _wallinfo_get(wall_element,
                                              ch_wall_stress_ID[i - 1, :])
            if np.size(select_wall) == 0:
                print('mgt記述の注意: wall_stressの壁(レベル%g, ID%g)は'
                      'mgtの壁要素に見つからないためスキップしました'
                      % (ch_wall_stress_ID[i - 1, 0],
                         ch_wall_stress_ID[i - 1, 1]))
                continue
            select_wall = np.atleast_2d(select_wall)[0]
            RCbar_arrange, _rci = _wallinfo_get(RCwalls,
                                                ch_wall_stress_ID[i - 1, :])
            RCbar_arrange = (np.atleast_2d(RCbar_arrange)[0]
                             if np.size(RCbar_arrange) else np.array([]))

            # 断面(厚さ)情報取得
            thick_ID = select_wall[7]
            thick_index = int(find_index(thick_nos_arr, thick_ID))
            element_thick = np.array(
                np.atleast_2d(thickness_wall)[thick_index, 0:2], copy=True)
            thick_no2 = element_thick[0]
            element_thick = float(element_thick[1])

            # 壁低減率の取得
            if wall_r is None or np.size(wall_r) == 0:
                reduction = np.array([])
            else:
                reduction, _ri = _wallinfo_get(
                    np.atleast_2d(np.asarray(wall_r, dtype=float)),
                    ch_wall_stress_ID[i - 1, :])

            # 材料情報取得
            material_ID = select_wall[6]
            material_index = find_index(material[:, 0], material_ID)
            ele_material = material[material_index, :]

            # 壁サイズの設定
            n1 = find_index(node[:, 0], select_wall[2])
            n2 = find_index(node[:, 0], select_wall[3])
            n3 = find_index(node[:, 0], select_wall[4])
            wall_width = float(np.linalg.norm(node[n1, 1:4] - node[n2, 1:4]))
            wall_height = float(np.linalg.norm(node[n2, 1:4] - node[n3, 1:4]))

            stress_w = ch_wall_stress[2 * i - 2:2 * i, :]
            timecase = ch_wall_stress_ID[i - 1, 2]
            _lc_w = max(len(x) for x in LCNAME)
            LOAS_CASE_NAME = LCNAME[int(find_index(
                ch_load_case_no_wall, timecase))].ljust(_lc_w)
            row = num_section + thick_index
            ratio_one = np.zeros((2, 4))
            if ele_material[1] in (1, 2):
                pass  # 原典: 鉄骨壁の検定 (空実装)
            elif ele_material[1] == 3:  # RC壁の検定
                if timecase == 1:
                    QL = stress_w[:, 1:3]
                else:
                    base = (ch_wall_stress.shape[0]
                            // ch_load_case_no_wall.shape[0])
                    QL_ID = np.array([(r % base) for r in
                                      range(2 * i - 1, 2 * i + 1)])
                    for iq in range(2):
                        if QL_ID[iq] == 0:
                            QL_ID[iq] = base
                    QL = ch_wall_stress[QL_ID - 1][:, 1:3]
                (ratio_one, mr, mtext) = RCwall_ratio_analysis(
                    element_thick, wall_width, wall_height, stress_w,
                    timecase, ele_material[1:3], select_wall[0:2],
                    thick_no2, maxratios[load_no][row, 1:3],
                    maxratios_text[row][load_no], QL, qup_wall,
                    walldesign_index_rc, select_wall, RCbar_arrange,
                    reduction, wall_cover, LOAS_CASE_NAME)
                maxratios[load_no][row, 1:3] = mr
                maxratios_text[row][load_no] = mtext
                UNVERIFIED_NOTE['hit'].add('RC壁要素(板壁)')
            # NOTE: 原典はSRC(4)/CFT(5)は「検定(未設定)」で素通し
            wall_ratio[2 * i - 2:2 * i, 3:7] = np.asarray(ratio_one,
                                                          dtype=float)

    # =======================================================================
    # トラス要素の断面検定
    # =======================================================================
    if _empty(ch_truss_stress):
        truss_ratio = np.zeros((0, 4))
    else:
        num_truss = ch_truss_stress.shape[0]
        truss_ratio = np.zeros((num_truss, 4))

        element_no_vec = ch_truss_stress[:, 0]
        material_no_vec = materialget(element_no_vec, element)
        material_index_vec = np.atleast_1d(
            find_index(material[:, 0], material_no_vec))
        ele_material_t = material[material_index_vec, :]

        # 断面情報一括取得 (ボルト欠損・応力倍率設定)
        Lminus = []
        Truss_times = []
        sec_no_list = element[np.atleast_1d(
            find_index(element[:, 0], element_no_vec)), 2]
        sec_no_list = doublecheck(sec_no_list)
        truss_bolt_loss = truss_bolt_loss or {}
        truss_times = truss_times or {}
        for i in range(sec_no_list.shape[0]):
            sec_no_i = sec_no_list[i]
            first_ele = element[find_index(element[:, 2], sec_no_i), 0]
            ss = sectionget(first_ele, element, sections, limit_sec_no)
            ss = ss[1:]
            if sec_no_i < limit_sec_no:
                code = ss[len(ss) - 2]
                if code == 2 or code == 5 or code == 9:
                    # ボルト欠損: 既定「欠損なし」= 0
                    Lminus.append([sec_no_i,
                                   float(truss_bolt_loss.get(int(sec_no_i), 0))])
                # 応力倍率: 既定 1.00
                Truss_times.append([sec_no_i,
                                    float(truss_times.get(int(sec_no_i), 1.0))])
        Lminus = np.asarray(Lminus, dtype=float) if Lminus else None
        Truss_times = (np.asarray(Truss_times, dtype=float)
                       if Truss_times else np.zeros((0, 2)))

        # 木造トラス用: 材料から壁倍率を規定 (MIDASratioplotW_fig.m の
        # w_panel_co listdlg → 引数 w_panel_co dict {断面番号: 壁倍率})
        w_panel_co = w_panel_co or {}
        w_panel_coefficient = (
            np.asarray([[float(k), float(v)] for k, v in w_panel_co.items()],
                       dtype=float)
            if w_panel_co else np.zeros((0, 2)))

        for i in range(num_truss):
            truss_ratio[i, 0:2] = ch_truss_stress[i, 0:2]
            load_case = ch_truss_stress[i, 1]
            # 元コード: 梁のケースリストで引く (verbatim)
            load_no = find_index(ch_load_case_no_beam, load_case)
            LOAS_CASE_NAME = LCNAME[load_no].ljust(
                max(len(x) for x in LCNAME))  # MATLAB char行のパディング再現

            element_no = ch_truss_stress[i, 0]
            # 応力ファイルにあるがmgtに無い要素は検定不能のためスキップ
            if find_index(element[:, 0], element_no) == -1:
                if int(element_no) not in missing_ele_warned:
                    missing_ele_warned.add(int(element_no))
                    print('mgt記述の注意: 応力ファイルの要素%dはmgtファイルに'
                          '存在しないため検定をスキップしました'
                          % int(element_no))
                continue
            sectionsize = sectionget(element_no, element, sections, limit_sec_no)
            section_no = sectionsize[0]
            pick_section_name = pick_text(
                section_name[find_index(A_section_no, section_no)], '_', -1)

            if section_no < limit_sec_no:
                sectionsize = sectionsize[1:]
                section_index = find_index(section_nos, section_no)
                stress = ch_truss_stress[i, [2, 3]]
                timecase = ch_truss_stress[i, 1]
                nodei_index, nodej_index = node_index_get(element_no, element,
                                                          node)
                ele_length = float(np.linalg.norm(
                    node[nodei_index, 1:4] - node[nodej_index, 1:4]))
                # 元コード: cb_judge は梁ループの残置変数のまま渡される
                buck_length = getbuck(ele_length, element_no, lky, lkz, lb,
                                      select_length, cb_judge,
                                      H_beam_RCsupport, stress)

                if ele_material_t[i, 1] == 1 or ele_material_t[i, 1] == 2:
                    (ratio_one, mr, mtext) = S_truss_ratio_analysis(
                        sectionsize, buck_length, stress, timecase,
                        ele_material_t[i, 1:4], element_no,
                        maxratios[load_no][section_index, 1:3],
                        maxratios_text[section_index][load_no],
                        section_no, Lminus, Truss_times, LOAS_CASE_NAME,
                        pick_section_name)
                    maxratios[load_no][section_index, 1:3] = mr
                    maxratios_text[section_index][load_no] = mtext
                    truss_ratio[i, 2:4] = np.asarray(ratio_one).ravel()[:2]
                elif ele_material_t[i, 1] == 10:
                    # 木造筋かい・壁面の検定 (MIDASratioplotW_fig.m:
                    # W_truss_ratio_analysis, Xブレース置換前提)
                    if (w_panel_coefficient.shape[0] == 0
                            or find_index(w_panel_coefficient[:, 0],
                                          section_no) == -1):
                        raise RuntimeError(
                            '木造トラス要素 (要素%d, 断面番号%d) の壁倍率が'
                            '未指定です。w_panel_co={断面番号: 壁倍率} で'
                            '指定してください (MATLAB版は w_panel_co.m の'
                            'listdlg で対話指定)。'
                            % (int(element_no), int(section_no)))
                    (ratio_one, mr, mtext) = W_truss_ratio_analysis(
                        stress, timecase, element_no,
                        maxratios[load_no][section_index, 1:3],
                        maxratios_text[section_index][load_no],
                        section_no, section_index, w_panel_coefficient,
                        element, node)
                    maxratios[load_no][section_index, 1:3] = mr
                    maxratios_text[section_index][load_no] = mtext
                    truss_ratio[i, 2:4] = np.asarray(ratio_one).ravel()[:2]
                else:
                    # MATLAB版は disp+stop。本移植はスコープ外としてスキップ記録
                    skipped.append({
                        'element_no': int(element_no),
                        'material_no': int(material_no_vec[i]
                                           if np.ndim(material_no_vec) else
                                           material_no_vec),
                        'material_type': int(ele_material_t[i, 1]),
                        'material_type_name': _MAT_TYPE_NAME.get(
                            int(ele_material_t[i, 1]),
                            str(int(ele_material_t[i, 1]))),
                        'section_no': int(section_no),
                        'load_case': int(load_case),
                        'kind': 'truss'})

    # ---- 結果詰め込み ----
    result = CheckResult()
    result.beam_ratio = beam_ratio
    result.truss_ratio = truss_ratio
    result.maxratios = maxratios
    result.maxratios_text = maxratios_text
    result.lb_table = lb_table
    result.section_nos = section_nos
    result.section_names = [
        section_name[find_index(A_section_no, sn)] for sn in section_nos]
    result.load_case_no = ch_load_case_no_beam
    result.LCNAME = LCNAME
    result.raw_case_no = _raw_case_no
    result.L_Lcase = np.asarray(L_Lcase, dtype=float).ravel()
    result.H_Lcase = np.asarray(H_Lcase, dtype=float).ravel()
    result.S_Lcase = np.asarray(S_Lcase, dtype=float).ravel()
    result.pm_direction = int(pm_direction)
    result.check_case_num = check_case_num
    result.skipped = skipped
    result.section_t_tex = section_t_tex
    result.load_direction = load_direction
    result.element_info = element
    result.plate_ratio = plate_ratio
    result.wall_ratio = wall_ratio
    result.thick_nos = np.asarray(thick_nos, dtype=float).ravel()
    result.num_section = num_section
    result.num_thick_wall = int(num_thick_wall)
    result.thickness = (np.atleast_2d(np.asarray(thickness, dtype=float))
                        if np.size(thickness) else np.zeros((0, 3)))
    if UNVERIFIED_NOTE['hit']:
        print('注記: 次の検定エンジンは移植済みですがMATLAB実行結果との'
              '照合例がまだありません(未検証): '
              + '、'.join(sorted(UNVERIFIED_NOTE['hit']))
              + '。実務利用の前に照合用データ(mgt+応力+MATLAB結果)での'
                '検証を推奨します。')
        if 'CFT' in UNVERIFIED_NOTE['hit']:
            print('注記: CFTの座屈長さは原典(MATLAB)どおり節点座標系の値を'
                  'そのままmm扱いで許容応力度式に渡しています(m系モデルでは'
                  '実質的に短柱扱い=座屈低減が効かず非安全側)。細長い柱では'
                  '注意してください。')
            print('注記: 角形CFT(EBC)は原典(MATLAB)が動作しない経路のため '
                  '(SA_CFTSBcolumn_AIJのBBOX呼び出し引数不足)、tx=ty=t を'
                  '補完して算定しています(NOTE SB_BBOX_T4)。')
    if TSR_SIGN_FIX['hit']:
        # ユーザー承認による原典(MATLAB)バグ修正 2026-07-08
        print('注記: 引張材(丸鋼)に圧縮が生じたため、負となる検定値の符号を'
              '正に修正して評価しています (原典(MATLAB)バグ修正 2026-07-08。'
              'MATLAB版実行結果と差が出る場合あり)。')
    if RC4_YOSE_FIX['hit']:
        # ユーザー承認による原典(MATLAB)バグ修正 2026-07-11
        print('注記: RC角柱の主筋を等間隔配筋(SA_RC4_HMD)で算定しています '
              '(原典(MATLAB)は恒真バグで常によせ筋版(端から76.5mm)を使用。'
              'ユーザー承認による原典バグ修正 2026-07-11。せい方向4段以上の'
              '柱はMATLAB版実行結果と差が出ます)。')
    if RCSR_NMZ_FIX['hit']:
        # ユーザー承認による原典(MATLAB)バグ修正 2026-07-12
        print('注記: RC丸柱の弱軸NM検定を曲げ項と軸力項の大きい方で評価して'
              'います (原典は軸力項が無視される既知バグ。ユーザー承認による'
              '原典バグ修正 2026-07-12。MATLAB版実行結果と差が出る場合あり)。')
    if SRC_CHECK_NOTE['hit']:
        print('注記: SRC部材を検定しました。SRC検定はKCUA物件(RH2T十字柱+'
              'RHB梁)でMATLAB結果と全数一致を確認済(2026-07-09)。'
              '鋼管被覆円柱等 当該照合に無い経路は未照合。')
    return result


# ===========================================================================
# Excel出力 (要素別_最大検定値)
# ===========================================================================

def export_maxratio_xlsx(result, out_path):
    """検定結果を「要素別_最大検定値」形式でExcel出力する (openpyxl).

    シート構成:
      要素別_最大検定値: 要素番号ごとに各検定ケースの最大検定値
                          (i/中央/j端×[cb,sz,sy,combi] の最大) と全ケース最大
      断面別最大        : ratio_260630.xlsm 'ratio'シート互換
                          (区分/ＩＤ/name/ケースごとの[最大要素番号, 最大検定値])
      beam_ratio        : 生の beam_ratio (3行/要素×ケース)
    """
    import openpyxl

    wb = openpyxl.Workbook()

    # ---- 要素別_最大検定値 ----
    ws = wb.active
    ws.title = '要素別_最大検定値'
    ncase = len(result.LCNAME)
    header = ['要素番号', '断面番号', '断面名']
    header += [str(name).strip() for name in result.LCNAME]
    header += ['最大検定値', '最大ケース']
    ws.append(header)

    br = result.beam_ratio
    sec_names = {int(sn): str(nm).strip()
                 for sn, nm in zip(result.section_nos, result.section_names)}
    # 要素→断面
    ele_case_max = {}
    order = []
    for r0 in range(0, br.shape[0], 3):
        ele = int(br[r0, 0])
        case = br[r0, 1]
        cidx = find_index(result.load_case_no, case)
        v = float(np.max(br[r0:r0 + 3, 2:6]))
        if ele not in ele_case_max:
            ele_case_max[ele] = [None] * ncase
            order.append(ele)
        ele_case_max[ele][cidx] = v
    for i in range(result.truss_ratio.shape[0]):
        ele = int(result.truss_ratio[i, 0])
        case = result.truss_ratio[i, 1]
        cidx = find_index(result.load_case_no, case)
        v = float(np.max(result.truss_ratio[i, 2:4]))
        if ele not in ele_case_max:
            ele_case_max[ele] = [None] * ncase
            order.append(ele)
        ele_case_max[ele][cidx] = v

    ele_sec = {}
    ei = result.element_info
    if ei is not None and np.asarray(ei).size:
        for r in range(np.asarray(ei).shape[0]):
            ele_sec[int(ei[r, 0])] = int(ei[r, 2])
    for ele in order:
        vals = ele_case_max[ele]
        valid = [v for v in vals if v is not None]
        vmax = max(valid) if valid else 0.0
        imax = vals.index(vmax) if valid else -1
        sec = ele_sec.get(ele)
        row = [ele, sec, sec_names.get(sec, '') if sec is not None else '']
        row += [v for v in vals]
        row += [vmax, str(result.LCNAME[imax]).strip() if imax >= 0 else '']
        ws.append(row)

    # ---- 断面別最大 (ratio_260630.xlsm 'ratio'シート互換) ----
    ws2 = wb.create_sheet('断面別最大')
    hdr = ['区分', 'ＩＤ', 'name']
    for name in result.LCNAME:
        hdr += [str(name).strip(), None]
    ws2.append(hdr)
    for si in range(result.section_nos.shape[0]):
        row = [None, int(result.section_nos[si]),
               sec_names.get(int(result.section_nos[si]), '')]
        for ci in range(min(len(result.maxratios), len(result.LCNAME))):
            if si < result.maxratios[ci].shape[0]:
                row += [int(result.maxratios[ci][si, 1]),
                        float(result.maxratios[ci][si, 2])]
            else:
                row += [None, None]
        ws2.append(row)

    # ---- 断面別最大: 壁厚/板厚の行 (板・壁要素の検定; 2026-07-13) ----
    _thick_nos = np.asarray(getattr(result, 'thick_nos',
                                    np.zeros(0)), dtype=float).ravel()
    _n_sec = int(getattr(result, 'num_section',
                         result.section_nos.shape[0]))
    _n_tw = int(getattr(result, 'num_thick_wall', 0))
    _ttbl = np.atleast_2d(np.asarray(getattr(result, 'thickness',
                                             np.zeros((0, 3))), dtype=float))
    for ti in range(_thick_nos.shape[0]):
        row_i = _n_sec + ti
        tid = _thick_nos[ti]
        name = ''
        if _ttbl.size:
            _tidx = find_index(_ttbl[:, 0], tid)
            _tidx = int(np.atleast_1d(_tidx)[0])
            if _tidx != -1:
                name = 't=%gmm' % _ttbl[_tidx, 1]
        row = ['壁厚' if ti < _n_tw else '板厚', int(tid), name]
        for ci in range(min(len(result.maxratios), len(result.LCNAME))):
            m = np.atleast_2d(result.maxratios[ci])
            if row_i < m.shape[0]:
                row += [int(m[row_i, 1]), float(m[row_i, 2])]
            else:
                row += [None, None]
        ws2.append(row)

    # ---- beam_ratio 生データ ----
    ws3 = wb.create_sheet('beam_ratio')
    ws3.append(['要素番号', '荷重ケース', 'ratio_cb', 'ratio_sz', 'ratio_sy',
                'ratio_combi'])
    for r in range(br.shape[0]):
        ws3.append([int(br[r, 0]), int(br[r, 1]), br[r, 2], br[r, 3],
                    br[r, 4], br[r, 5]])
    if result.truss_ratio.shape[0]:
        ws4 = wb.create_sheet('truss_ratio')
        ws4.append(['要素番号', '荷重ケース', 'ratio_1', 'ratio_2'])
        for r in range(result.truss_ratio.shape[0]):
            ws4.append([int(result.truss_ratio[r, 0]),
                        int(result.truss_ratio[r, 1]),
                        result.truss_ratio[r, 2], result.truss_ratio[r, 3]])

    wb.save(out_path)
    return out_path
# ===========================================================================
# Excel出力 (要素別全結果: beam_ratio / truss_ratio 行列の生値)
# ===========================================================================

def export_full_ratio_xlsx(result, out_path):
    """検定結果の要素別全結果 (beam_ratio/truss_ratio 行列の生値) をExcel出力.

    MATLAB版でワークスペースの beam_ratio / truss_ratio 行列をExcelへ
    コピペして詳細検討していた作業の代替。行の並び・数値は行列そのまま
    (丸めない) とし、要素番号・ケース名・断面情報の列を付加する。

    シート構成:
      beam_ratio : beam_ratio の1行=1行
                   (要素あたり i端/中央/j端 の3行×検定ケース, 検定比4列)
      truss_ratio: truss_ratio の1行=1行 (1行に i端/j端 の検定比2列)
      凡例       : 列の意味・検定ケース一覧・注意事項
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    case_nos = np.atleast_1d(np.asarray(result.load_case_no, dtype=float))
    case_names = [str(n).strip() for n in result.LCNAME]

    def _case_name(case):
        ci = find_index(case_nos, case)
        ci = int(np.atleast_1d(ci)[0]) if np.atleast_1d(ci).size else -1
        return case_names[ci] if 0 <= ci < len(case_names) else ''

    sec_names = {int(sn): str(nm).strip()
                 for sn, nm in zip(result.section_nos, result.section_names)}
    ele_sec = {}
    ei = np.asarray(result.element_info)
    if ei.size:
        for r in range(ei.shape[0]):
            ele_sec[int(ei[r, 0])] = int(ei[r, 2])

    bold = Font(bold=True)

    def _setup_sheet(ws, header, widths):
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = bold
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'C2'  # 先頭2列(要素番号/検定ケース番号)+ヘッダ行を凍結

    wb = openpyxl.Workbook()

    # ---- beam_ratio (生の並びを保持: 3行/要素×検定ケース) ----
    ws = wb.active
    ws.title = 'beam_ratio'
    _setup_sheet(ws, ['要素番号', '検定ケース番号', 'ケース名', '位置',
                      '断面番号', '断面名', '検定比(曲げ+軸力)', '検定比(せん断強軸)',
                      '検定比(せん断弱軸)', '検定比(組合せ応力)'],
                 [10, 14, 16, 8, 10, 28, 13, 13, 13, 13])
    pos_label = ('i端', '中央', 'j端')
    br = np.asarray(result.beam_ratio, dtype=float)
    for r in range(br.shape[0]):
        ele = int(br[r, 0])
        sec = ele_sec.get(ele)
        ws.append([ele, int(br[r, 1]), _case_name(br[r, 1]),
                   pos_label[r % 3], sec,
                   sec_names.get(sec, '') if sec is not None else '',
                   float(br[r, 2]), float(br[r, 3]), float(br[r, 4]),
                   float(br[r, 5])])
    ws.auto_filter.ref = 'A1:J%d' % (br.shape[0] + 1)

    # ---- truss_ratio (生の並びを保持: 1行/要素×検定ケース, i端/j端の2列) ----
    tr = np.asarray(result.truss_ratio, dtype=float)
    ws2 = wb.create_sheet('truss_ratio')
    _setup_sheet(ws2, ['要素番号', '検定ケース番号', 'ケース名',
                       '断面番号', '断面名', '検定比(i端)', '検定比(j端)'],
                 [10, 14, 16, 10, 28, 13, 13])
    for r in range(tr.shape[0]):
        ele = int(tr[r, 0])
        sec = ele_sec.get(ele)
        ws2.append([ele, int(tr[r, 1]), _case_name(tr[r, 1]), sec,
                    sec_names.get(sec, '') if sec is not None else '',
                    float(tr[r, 2]), float(tr[r, 3])])
    ws2.auto_filter.ref = 'A1:G%d' % (tr.shape[0] + 1)

    # ---- plate_ratio (板要素の検定; 2026-07-13 未検証) ----
    pr = np.atleast_2d(np.asarray(getattr(result, 'plate_ratio',
                                          np.zeros((0, 4))), dtype=float))
    if pr.size:
        wsp = wb.create_sheet('plate_ratio')
        _setup_sheet(wsp, ['板要素番号', '検定ケース番号', 'ケース名',
                           '節点番号', '検定比'],
                     [12, 14, 16, 10, 13])
        for r in range(pr.shape[0]):
            wsp.append([int(pr[r, 0]), int(pr[r, 1]), _case_name(pr[r, 1]),
                        int(pr[r, 2]), float(pr[r, 3])])
        wsp.auto_filter.ref = 'A1:E%d' % (pr.shape[0] + 1)

    # ---- wall_ratio (壁要素の検定; 2026-07-13 未検証) ----
    wr = np.atleast_2d(np.asarray(getattr(result, 'wall_ratio',
                                          np.zeros((0, 7))), dtype=float))
    if wr.size:
        wsw = wb.create_sheet('wall_ratio')
        _setup_sheet(wsw, ['ベースレベル', '壁ID', '検定ケース番号',
                           'ケース名', '応力行', '検定比(NMy)', '検定比(NMz)',
                           '検定比(Qz)', '検定比(Qy)'],
                     [12, 8, 14, 16, 8, 13, 13, 13, 13])
        for r in range(wr.shape[0]):
            wsw.append([float(wr[r, 0]), int(wr[r, 1]), int(wr[r, 2]),
                        _case_name(wr[r, 2]), (r % 2) + 1,
                        float(wr[r, 3]), float(wr[r, 4]),
                        float(wr[r, 5]), float(wr[r, 6])])
        wsw.auto_filter.ref = 'A1:I%d' % (wr.shape[0] + 1)

    # ---- 凡例 ----
    ws3 = wb.create_sheet('凡例')
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 90

    def _title(text):
        ws3.append([text])
        ws3.cell(row=ws3.max_row, column=1).font = bold

    _title('列の意味 (beam_ratio / truss_ratio 共通)')
    for k, v in (
            ('要素番号', 'mgtファイルの要素番号'),
            ('検定ケース番号', '検定用組合せケースの番号 (長期=1, 組合せ'
             'ケース=長期ケース番号×10×加力方向+水平ケース番号 等)'),
            ('ケース名', '検定ケース名 (LCNAME)'),
            ('断面番号', '要素に割り当てられた断面番号 (mgtのELEMENTより)'),
            ('断面名', '断面番号に対応する断面名 (ダミー材は空欄)'),
    ):
        ws3.append([k, v])
    ws3.append([])
    _title('beam_ratio シート固有の列 (要素あたり i端/中央/j端 の3行)')
    for k, v in (
            ('位置', '検定位置 (i端/中央/j端)'),
            ('検定比(曲げ+軸力)', '曲げ+軸力の検定比 (無次元, 検定応力度/許容応力度)'),
            ('検定比(せん断強軸)', 'せん断(強軸方向)の検定比 (無次元)'),
            ('検定比(せん断弱軸)', 'せん断(弱軸方向)の検定比 (無次元)'),
            ('検定比(組合せ応力)', '組合せ応力の検定比 (無次元)'),
    ):
        ws3.append([k, v])
    ws3.append([])
    _title('RC要素の列構成 (material type 3, RC_ratio_analysis)')
    for k, v in (
            ('RC梁', '検定比(曲げ+軸力)=曲げ検定比、検定比(せん断強軸)='
             'せん断検定比。検定比(せん断弱軸)・(組合せ応力) の列は常に0。'),
            ('RC壁柱', '検定比(曲げ+軸力)=NMy(軸力+面外曲げ)、'
             '検定比(せん断強軸)=NMz(軸力+面内曲げ)、'
             '検定比(せん断弱軸)=Qz(面外せん断)、'
             '検定比(組合せ応力)=Qy(面内せん断)。中央行は常に0 (i端/j端のみ検定)。'),
    ):
        ws3.append([k, v])
    ws3.append([])
    _title('truss_ratio シート固有の列 (要素あたり1行)')
    for k, v in (
            ('検定比(i端)', 'i端軸力による引張/圧縮の検定比 (無次元)'),
            ('検定比(j端)', 'j端軸力による引張/圧縮の検定比 (無次元)'),
    ):
        ws3.append([k, v])
    ws3.append([])
    _title('検定ケース一覧')
    ws3.append(['ケース番号', 'ケース名', '種別'])
    ws3.cell(row=ws3.max_row, column=1).font = bold
    ws3.cell(row=ws3.max_row, column=2).font = bold
    ws3.cell(row=ws3.max_row, column=3).font = bold
    for i in range(case_nos.shape[0]):
        nm = case_names[i] if i < len(case_names) else ''
        kind = '長期' if case_nos[i] < 10 else '短期'
        ws3.append([int(case_nos[i]), nm, kind])
    ws3.append([])
    _title('注意事項')
    for note in (
            '検定値は生値のまま出力している (丸めなし)。',
            '木造要素 (material type 10) も同じ4列構成で出力される。'
            '検定比(曲げ+軸力)=二軸曲げ+軸力、検定比(組合せ応力)は木造では'
            '常に0 (W_ratio_analysis の出力列構成)。',
            'RC要素は上記「RC要素の列構成」の並びで出力される。',
            'PC/CFT/RC杭/RM(補強組積造)の要素も検定値が出力されるが、'
            'これらの経路は検証データ(MATLAB照合)が無いため未検証 '
            '(2026-07-13移植)。値の妥当性は設計者が確認すること。',
            'plate_ratio(板要素)・wall_ratio(壁要素)シートは plate_stress/'
            'wall_stress 入力時のみ出力される。これらの経路も未検証。'
            'wall_ratio の検定比列の構成はRC壁柱と同じ '
            '(NMy=軸力+面外曲げ, NMz=軸力+面内曲げ, Qz=面外せん断, '
            'Qy=面内せん断)。応力行=wall_stressの2行1組の行順。',
            'SRC要素(material type 5): KCUA物件(RH2T十字柱+RHB梁)で'
            'MATLAB結果と全数一致を確認済(2026-07-09)。鋼管被覆円柱等'
            '当該照合に無いSRC経路は未照合。',
            'ダミー材(limit_sec_no以上の断面番号)・mgtに存在しない要素も'
            '検定値0で出力される。',
            '行の並びは beam_ratio / truss_ratio 行列 (MATLAB版ワーク'
            'スペース変数) と同一。',
            '引張材(丸鋼)が圧縮となる場合の検定値の符号を修正済み '
            '(負→正。ユーザー承認による原典(MATLAB)バグ修正 2026-07-08。'
            'MATLAB版実行結果と差が出る場合あり)。',
            'RC角柱の主筋は等間隔配筋で算定 (原典(MATLAB)は恒真バグで常に'
            'よせ筋版(端から76.5mm)。ユーザー承認による原典バグ修正 '
            '2026-07-11。せい方向4段以上の柱はMATLAB版と差が出る場合あり)。',
            'RC丸柱の弱軸NM検定は曲げ項と軸力項の大きい方で評価 (原典は'
            '軸力項が無視される既知バグ。ユーザー承認による原典バグ修正 '
            '2026-07-12。MATLAB版と差が出る場合あり)。',
    ):
        ws3.append(['', note])

    # openpyxl は数値セルを "%.16g" で書き出すため float64 の最下位桁
    # (1ULP) が失われることがある。生値の完全往復のため、保存時のみ
    # float の書式を repr (最短往復表現, 17有効桁まで) に差し替える。
    import openpyxl.cell._writer as _cell_writer
    _orig_safe_string = _cell_writer.safe_string

    def _roundtrip_safe_string(value):
        if (isinstance(value, float)
                and not (math.isnan(value) or math.isinf(value))):
            return repr(value)
        return _orig_safe_string(value)

    _cell_writer.safe_string = _roundtrip_safe_string
    try:
        wb.save(out_path)
    finally:
        _cell_writer.safe_string = _orig_safe_string
    return out_path
