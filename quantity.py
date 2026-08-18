# -*- coding: utf-8 -*-
"""建築材料数量集計 (RC梁・柱・壁): 配筋 x 部材長 から鉄筋質量・コンクリート体積を算出する.

新規実装 (MATLAB原典なし)。mgt パーサ (mgtopen_element/node/material/section/
RCbeam/RCcolumn/wall/RCwall/thickness) をそのまま利用し、数量集計ロジックを
追加する。断面検定 (rc_check.py 等) とは独立で、検定の合否判定は行わない。

対象: RC梁・RC柱・RC壁・RCスラブ(板要素)。杭は対象外 (今後拡張の可能性あり)。

主な前提 (概算のための割り切り。詳細な配筋要領図の本数・定尺とは一致しない):
- 梁の主筋・あばら筋は i端/中央/j端の3断面のデータをそれぞれ全長の1/3に
  適用する (mgtの配筋データ自体がこの3断面区分で表現されているため)。
  あばら筋の「本数」は各断面(1/3区間)内の本数として mgt にそのまま
  入力されている値を用いる。
- 柱のフープ本数(セット数)は 部材長 / ピッチ + 1 (端数切捨て) で概算する
  (継手・重ね代・上下端の増し打ち・フック長さは考慮しない)。Fy方向本数は
  D方向(成)の脚、Fz方向本数はb方向(幅)の脚として長さを計算する
  (MIDASの実際の脚配置とは一致しない可能性があるため要確認)。
  脚1本の長さは芯かぶり(既定40mm、UIで変更可)を控除したコア長さのみ。
  梁のあばら筋は b・D 双方の周長式 (2x((b-2c)+(D-2c))) を用いる。
- 壁の縦筋・横筋本数は (幅または高さ/ピッチ+1) で概算し、既定で複配筋
  (x2、UIで単配筋に変更可)。端部補強筋は *REBAR-WALL の本数をそのまま
  高さ方向に適用する。断面名に'W'を含み配筋定義(*REBAR-WALL)を持たない
  線材(梁・柱要素形式)の壁パネルは、断面検定タブと同様に断面符号ごとに
  ユーザーが配筋を入力する (rc_wall_line_quantities)。
- スラブ(RC板要素、PLATE/PLSTRS/PLSTRN/AXISYM)は mgt に配筋データが
  無いため、厚みIDごとにユーザーがX方向筋・Y方向筋(径・ピッチ)と
  上端筋のみ/上下端筋を入力する (rc_slab_quantities)。多数のメッシュ要素で
  構成される連続体として扱い、1要素ごとに本数を丸めず「厚み区分の合計面積
  ÷ピッチ」で延長を概算する (壁・柱のような部材単位の+1丸めは行わない)。
- 鉄筋質量は JIS公称断面積(bar_table.json) x 鋼材密度(7850kg/m3) x 長さ。
"""

import json
import math
import os

import numpy as np

from .util import find_index
from .mgt import (mgtopen_node, mgtopen_element, mgtopen_material,
                  mgtopen_RCbeam, mgtopen_RCcolumn, mgtopen_wall,
                  mgtopen_RCwall, mgtopen_thickness, mgtopen_thickness_name,
                  mgtopen_plate)
from .section import mgtopen_section

_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
STEEL_DENSITY = 7.85e-3  # kg / (mm2 * m)  (= 7850 kg/m3)

with open(os.path.join(_DATA_DIR, 'bar_table.json'), encoding='utf-8') as _f:
    _BAR = json.load(_f)
_BAR_AREA1 = {int(d): float(a[0])
              for d, a in zip(_BAR['diameters'], _BAR['areas'])}


def bar_area(dia):
    """公称直径(mm)から1本あたりの公称断面積(mm2)を返す (bar_table.json)."""
    d = int(round(float(dia)))
    a = _BAR_AREA1.get(d)
    if a is None:
        raise ValueError('未対応の鉄筋径です: D%s' % dia)
    return a


def bar_weight_per_m(dia):
    """公称直径(mm)から単位長さ質量(kg/m)を返す (公称断面積x鋼材密度)."""
    return bar_area(dia) * STEEL_DENSITY


class Tally(dict):
    """{径(int): {'length': 延長(m), 'weight': 質量(kg)}} の集計ヘルパー."""

    def add(self, dia, length_m, count=1.0):
        if dia is None or length_m is None:
            return
        dia = float(dia)
        length_m = float(length_m)
        count = float(count)
        if dia <= 0 or length_m <= 0 or count <= 0:
            return
        d = int(round(dia))
        total_len = length_m * count
        row = self.setdefault(d, {'length': 0.0, 'weight': 0.0})
        row['length'] += total_len
        row['weight'] += total_len * bar_weight_per_m(d)

    def merge(self, other):
        for d, v in other.items():
            row = self.setdefault(d, {'length': 0.0, 'weight': 0.0})
            row['length'] += v['length']
            row['weight'] += v['weight']


def _leg_length_m(dim_m, cover_m, support_dia_mm):
    """あばら筋/フープ1脚分のコア長さ(芯かぶり控除後、m)."""
    return max(dim_m - 2.0 * cover_m - support_dia_mm / 1000.0, 0.0)


def _tie_length_m(b_m, D_m, cover_m, support_dia_mm):
    """閉鎖形あばら筋/フープ1本の概算全長(m) (芯かぶり控除後のコア周長のみ)."""
    return 2.0 * (_leg_length_m(b_m, cover_m, support_dia_mm)
                  + _leg_length_m(D_m, cover_m, support_dia_mm))


def _sb_table(sections):
    """sections[2] (中実角断面 [断面番号,D1,D2] 、単位m) を取り出す."""
    if len(sections) > 2 and sections[2] is not None and np.asarray(sections[2]).size:
        return np.atleast_2d(np.asarray(sections[2], dtype=float))
    return np.zeros((0, 3))


def _node_pos(node):
    node = np.atleast_2d(node)
    return {int(r[0]): (float(r[1]), float(r[2]), float(r[3]))
            for r in node} if node.size else {}


def _ele_length(pos, ni, nj):
    xi, yi, zi = pos[ni]
    xj, yj, zj = pos[nj]
    return math.sqrt((xi - xj) ** 2 + (yi - yj) ** 2 + (zi - zj) ** 2)


def _polygon_area(pts):
    """3辺以上の点列(3D、平面上)から多角形面積(m2相当)を返す (扇形三角形分割)."""
    if len(pts) < 3:
        return 0.0
    p0 = np.asarray(pts[0], dtype=float)
    area = 0.0
    for i in range(1, len(pts) - 1):
        v1 = np.asarray(pts[i], dtype=float) - p0
        v2 = np.asarray(pts[i + 1], dtype=float) - p0
        area += 0.5 * float(np.linalg.norm(np.cross(v1, v2)))
    return area


# ---------------------------------------------------------------------------
# RC梁
# ---------------------------------------------------------------------------

def rc_beam_quantities(mgt_path, cover_m=0.04):
    """RC梁 (*REBAR-BEAM のある断面) の主筋・あばら筋質量とコンクリート体積.

    戻り値: dict {'rebar': Tally, 'concrete_m3': float,
                 'rows': [要素別明細], 'skipped': [除外要素の理由]}
    """
    node = mgtopen_node(mgt_path)
    element = np.atleast_2d(mgtopen_element(mgt_path))
    material = np.atleast_2d(mgtopen_material(mgt_path))
    sb = _sb_table(mgtopen_section(mgt_path)[0])
    rcbeams = mgtopen_RCbeam(mgt_path)
    rcbeam_map = {int(sec): mat for sec, mat in rcbeams}
    pos = _node_pos(node)

    rebar = Tally()
    concrete_m3 = 0.0
    rows = []
    skipped = []

    for r in range(element.shape[0]):
        m_idx = find_index(material[:, 0], element[r, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        s_no = int(element[r, 2])
        if s_no not in rcbeam_map:
            continue
        ele_no = int(element[r, 0])
        ni, nj = int(element[r, 3]), int(element[r, 4])
        if ni not in pos or nj not in pos:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '節点が見つかりません'})
            continue
        L = _ele_length(pos, ni, nj)
        sbi = find_index(sb[:, 0], s_no) if sb.size else -1
        if sbi == -1:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '矩形断面寸法(中実角)が見つかりません'})
            continue
        D = float(sb[sbi, 1])
        b = float(sb[sbi, 2])
        ele_concrete = b * D * L
        concrete_m3 += ele_concrete

        mat = rcbeam_map[s_no]
        for zone in range(3):
            zL = L / 3.0
            row = mat[zone]
            rebar.add(row[4], zL, row[1])           # 上端1段目
            if row[0] == 2:
                rebar.add(row[5], zL, row[2])       # 上端2段目
            rebar.add(row[11], zL, row[8])          # 下端1段目
            if row[7] == 2:
                rebar.add(row[12], zL, row[9])      # 下端2段目
            if row[14] > 0 and row[16] > 0:
                tie_len = _tie_length_m(b, D, cover_m, row[14])
                rebar.add(row[14], tie_len, row[16])  # あばら筋

        rows.append({'ele': ele_no, 'sec_no': s_no, 'length': L,
                     'b': b, 'D': D, 'concrete_m3': ele_concrete})

    return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
            'skipped': skipped}


# ---------------------------------------------------------------------------
# RC柱
# ---------------------------------------------------------------------------

def rc_column_quantities(mgt_path, cover_m=0.04):
    """RC柱 (*REBAR-COLUMN のある断面) の主筋・フープ質量とコンクリート体積."""
    node = mgtopen_node(mgt_path)
    element = np.atleast_2d(mgtopen_element(mgt_path))
    material = np.atleast_2d(mgtopen_material(mgt_path))
    sb = _sb_table(mgtopen_section(mgt_path)[0])
    rccols = np.asarray(mgtopen_RCcolumn(mgt_path), dtype=float)
    col_map = {}
    if rccols.size:
        rccols = np.atleast_2d(rccols)
        for row in rccols:
            col_map[int(row[0])] = row
    pos = _node_pos(node)

    rebar = Tally()
    concrete_m3 = 0.0
    rows = []
    skipped = []

    for r in range(element.shape[0]):
        m_idx = find_index(material[:, 0], element[r, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        s_no = int(element[r, 2])
        if s_no not in col_map:
            continue
        ele_no = int(element[r, 0])
        ni, nj = int(element[r, 3]), int(element[r, 4])
        if ni not in pos or nj not in pos:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '節点が見つかりません'})
            continue
        L = _ele_length(pos, ni, nj)
        sbi = find_index(sb[:, 0], s_no) if sb.size else -1
        if sbi == -1:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '矩形断面寸法(中実角)が見つかりません'})
            continue
        D = float(sb[sbi, 1])
        b = float(sb[sbi, 2])
        ele_concrete = b * D * L
        concrete_m3 += ele_concrete

        c = col_map[s_no]
        total_bars, main_dia = c[2], c[3]
        hoop_dia, hoop_pitch_mm = c[6], c[7]
        fy_num, fz_num = c[8], c[9]
        rebar.add(main_dia, L, total_bars)
        if hoop_dia > 0 and hoop_pitch_mm > 0:
            n_sets = max(1, math.floor(L * 1000.0 / hoop_pitch_mm) + 1)
            leg_D = _leg_length_m(D, cover_m, hoop_dia)
            leg_b = _leg_length_m(b, cover_m, hoop_dia)
            if fy_num > 0:
                rebar.add(hoop_dia, leg_D, n_sets * fy_num)
            if fz_num > 0:
                rebar.add(hoop_dia, leg_b, n_sets * fz_num)

        rows.append({'ele': ele_no, 'sec_no': s_no, 'length': L,
                     'b': b, 'D': D, 'concrete_m3': ele_concrete})

    return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
            'skipped': skipped}


# ---------------------------------------------------------------------------
# RC壁
# ---------------------------------------------------------------------------

def wall_plate_thicknesses(mgt_path):
    """配筋(*REBAR-WALL)のある板要素壁パネルが使う厚みID一覧を返す.

    厚みごとに単配筋/複配筋をユーザーに指定してもらうための候補一覧
    (rc_wall_quantities の layer_by_thickness 入力用)。
    戻り値: [{'thick_id','thickness_mm','name','n_walls'}, ...] (厚みID昇順)
    """
    node = mgtopen_node(mgt_path)
    wall_element, _wall_base = mgtopen_wall(mgt_path, node)
    wall_element = np.asarray(wall_element, dtype=float)
    if not wall_element.size:
        return []
    thickness = np.asarray(mgtopen_thickness(mgt_path), dtype=float)
    thickness_names = mgtopen_thickness_name(mgt_path)
    RCwalls, _FL_name = mgtopen_RCwall(mgt_path, wall_element)
    RCwalls = np.asarray(RCwalls, dtype=float)
    rc_keys = set()
    if RCwalls.size:
        for row in np.atleast_2d(RCwalls):
            rc_keys.add((int(row[1]), round(float(row[0]), 6)))

    counts = {}
    for wr in np.atleast_2d(wall_element):
        key = (int(wr[1]), round(float(wr[0]), 6))
        if key not in rc_keys:
            continue
        tid = int(wr[7])
        counts[tid] = counts.get(tid, 0) + 1

    out = []
    for tid, n in sorted(counts.items()):
        mm = None
        if thickness.size:
            ti = find_index(np.atleast_2d(thickness)[:, 0], tid)
            if ti != -1:
                mm = float(np.atleast_2d(thickness)[ti, 1])
        out.append({'thick_id': tid, 'thickness_mm': mm,
                    'name': thickness_names.get(tid, ''), 'n_walls': n})
    return out


def rc_wall_quantities(mgt_path, cover_m=0.04, layer_by_thickness=None,
                       default_layer='double'):
    """RC壁 (*REBAR-WALL のある壁ID) の縦横筋・端部補強筋質量とコンクリート体積.

    layer_by_thickness: {厚みID: 'single'|'double'} 厚みごとの単配筋/複配筋
      指定 (断面検定タブと同様、板要素壁は厚みIDに配筋情報が紐づかないため
      ユーザー指定が必要)。未指定の厚みは default_layer を使う。
    """
    node = mgtopen_node(mgt_path)
    wall_element, _wall_base = mgtopen_wall(mgt_path, node)
    wall_element = np.asarray(wall_element, dtype=float)
    rebar = Tally()
    concrete_m3 = 0.0
    rows = []
    skipped = []
    if not wall_element.size:
        return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
                'skipped': skipped}

    layer_by_thickness = layer_by_thickness or {}
    thickness = np.asarray(mgtopen_thickness(mgt_path), dtype=float)
    RCwalls, _FL_name = mgtopen_RCwall(mgt_path, wall_element)
    RCwalls = np.asarray(RCwalls, dtype=float)
    pos = _node_pos(node)

    rc_map = {}
    if RCwalls.size:
        for row in np.atleast_2d(RCwalls):
            rc_map[(int(row[1]), round(float(row[0]), 6))] = row

    for wr in np.atleast_2d(wall_element):
        base_level, wall_id = float(wr[0]), int(wr[1])
        node_ids = [int(v) for v in wr[2:6] if int(v) != 0]
        pts = [pos[n] for n in node_ids if n in pos]
        if len(pts) < 3:
            skipped.append({'wall_id': wall_id,
                            'reason': '節点が3点未満のため形状を計算できません'})
            continue
        zs = [p[2] for p in pts]
        height = max(zs) - min(zs)
        dirv = np.array([wr[8], wr[9]], dtype=float)
        norm = np.linalg.norm(dirv)
        if norm < 1e-9 or height <= 0:
            skipped.append({'wall_id': wall_id,
                            'reason': '壁の方向ベクトルまたは高さが取得できません'})
            continue
        dirv = dirv / norm
        proj = [p[0] * dirv[0] + p[1] * dirv[1] for p in pts]
        width = max(proj) - min(proj)
        if width <= 0:
            skipped.append({'wall_id': wall_id, 'reason': '壁幅が0です'})
            continue

        thick_id = int(wr[7])
        t = 0.0
        if thickness.size:
            ti = find_index(np.atleast_2d(thickness)[:, 0], thick_id)
            if ti != -1:
                t = float(np.atleast_2d(thickness)[ti, 1]) / 1000.0
        ele_concrete = width * height * t
        concrete_m3 += ele_concrete

        key = (wall_id, round(base_level, 6))
        rc = rc_map.get(key)
        if rc is None:
            skipped.append({'wall_id': wall_id, 'base_level': base_level,
                            'reason': '配筋(*REBAR-WALL)が見つかりません'})
            rows.append({'wall_id': wall_id, 'base_level': base_level,
                        'thick_id': thick_id, 'width': width,
                        'height': height, 'thickness_m': t,
                        'concrete_m3': ele_concrete})
            continue

        layer_mult = 1.0 if layer_by_thickness.get(
            thick_id, default_layer) == 'single' else 2.0
        v_dia, v_pitch_mm = rc[2], rc[3]
        h_dia, h_pitch_mm = rc[4], rc[5]
        edge_dia, _edge_pitch_mm, edge_num = rc[6], rc[7], rc[8]
        if v_dia > 0 and v_pitch_mm > 0:
            n_v = max(1, math.floor(width * 1000.0 / v_pitch_mm) + 1)
            rebar.add(v_dia, height, n_v * layer_mult)
        if h_dia > 0 and h_pitch_mm > 0:
            n_h = max(1, math.floor(height * 1000.0 / h_pitch_mm) + 1)
            rebar.add(h_dia, width, n_h * layer_mult)
        if edge_dia > 0 and edge_num > 0:
            rebar.add(edge_dia, height, edge_num)

        rows.append({'wall_id': wall_id, 'base_level': base_level,
                     'thick_id': thick_id, 'width': width, 'height': height,
                     'thickness_m': t, 'concrete_m3': ele_concrete})

    return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
            'skipped': skipped}


# ---------------------------------------------------------------------------
# RC壁 (線材=梁・柱要素形式でモデル化された壁、断面名に'W'を含むもの)
# ---------------------------------------------------------------------------

def wall_line_candidate_sections(mgt_path, limit_sec_no=9000.0):
    """断面名に'W'を含み、配筋定義(*REBAR-BEAM/COLUMN/WALL)を持たない
    RC(材料type3)の中実角断面 (線材でモデル化された壁パネルの候補) を返す.

    戻り値: [{'no','name','prefix','t','D'}, ...]  (t=厚みmm, D=壁長mm)
    断面検定タブの「配筋の無いRC断面」入力対象と同一条件 (rc_check.py
    RCW_input 相当) から、断面名に'W'を含むものだけを抜き出したもの。
    """
    element = np.atleast_2d(mgtopen_element(mgt_path))
    material = np.atleast_2d(mgtopen_material(mgt_path))
    sections, section_no, section_name = mgtopen_section(mgt_path)
    sb = _sb_table(sections)
    section_no = np.asarray(section_no, dtype=float).ravel()
    rcbeams = mgtopen_RCbeam(mgt_path)
    rebar_secs = set(int(sec) for sec, _mat in rcbeams)
    rccols = np.asarray(mgtopen_RCcolumn(mgt_path), dtype=float)
    if rccols.size:
        rebar_secs |= set(int(v) for v in np.atleast_2d(rccols)[:, 0])

    candidates = {}
    for r in range(element.shape[0]):
        m_idx = find_index(material[:, 0], element[r, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        s_no = int(element[r, 2])
        if s_no >= limit_sec_no or s_no in rebar_secs or s_no in candidates:
            continue
        si = find_index(section_no, s_no)
        name = str(section_name[si]).strip() if si != -1 else str(s_no)
        if 'W' not in name.upper():
            continue
        sbi = find_index(sb[:, 0], s_no) if sb.size else -1
        if sbi == -1:
            continue
        t = round(float(sb[sbi, 1]) * 1000, 3)
        D = round(float(sb[sbi, 2]) * 1000, 3)
        prefix = name.split('_', 1)[0].strip()
        candidates[s_no] = {'no': s_no, 'name': name, 'prefix': prefix,
                            't': t, 'D': D}
    return sorted(candidates.values(), key=lambda d: d['no'])


def rc_wall_line_quantities(mgt_path, rebar_settings):
    """線材(梁・柱要素形式)壁パネルの縦横筋・端部補強筋質量とコンクリート体積.

    rebar_settings: {断面番号(int): {'v_dia','v_pitch','h_dia','h_pitch',
                     'layer'('single'|'double'), 'edge_num'}}
      断面検定タブの「配筋の無いRC断面」入力(壁)と同じ項目・意味で、
      断面番号ごとにユーザーが指定する。厚みt・壁長Dは断面(中実角)から、
      高さは要素長さ(節点間距離)から取得する。edge_num は端部補強筋の
      片側本数 (縦筋径と同径、両端で edge_num*2 本として加算)。
    """
    node = mgtopen_node(mgt_path)
    element = np.atleast_2d(mgtopen_element(mgt_path))
    material = np.atleast_2d(mgtopen_material(mgt_path))
    sb = _sb_table(mgtopen_section(mgt_path)[0])
    pos = _node_pos(node)

    rebar = Tally()
    concrete_m3 = 0.0
    rows = []
    skipped = []

    for r in range(element.shape[0]):
        m_idx = find_index(material[:, 0], element[r, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        s_no = int(element[r, 2])
        cfg = rebar_settings.get(s_no)
        if cfg is None:
            continue
        ele_no = int(element[r, 0])
        ni, nj = int(element[r, 3]), int(element[r, 4])
        if ni not in pos or nj not in pos:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '節点が見つかりません'})
            continue
        L = _ele_length(pos, ni, nj)  # 壁パネルの高さ
        sbi = find_index(sb[:, 0], s_no) if sb.size else -1
        if sbi == -1:
            skipped.append({'ele': ele_no, 'sec_no': s_no,
                            'reason': '矩形断面寸法(中実角)が見つかりません'})
            continue
        t = float(sb[sbi, 1])  # 厚み(m)
        D = float(sb[sbi, 2])  # 壁長(m)
        ele_concrete = t * D * L
        concrete_m3 += ele_concrete

        layer_mult = 1.0 if str(cfg.get('layer', 'double')) == 'single' else 2.0
        v_dia = float(cfg.get('v_dia', 0) or 0)
        v_pitch = float(cfg.get('v_pitch', 0) or 0)
        h_dia = float(cfg.get('h_dia', 0) or 0)
        h_pitch = float(cfg.get('h_pitch', 0) or 0)
        edge_num = float(cfg.get('edge_num', 0) or 0)
        if v_dia > 0 and v_pitch > 0:
            n_v = max(1, math.floor(D * 1000.0 / v_pitch) + 1)
            rebar.add(v_dia, L, n_v * layer_mult)
        if h_dia > 0 and h_pitch > 0:
            n_h = max(1, math.floor(L * 1000.0 / h_pitch) + 1)
            rebar.add(h_dia, D, n_h * layer_mult)
        if edge_num > 0 and v_dia > 0:
            rebar.add(v_dia, L, edge_num * 2)  # 両端 (片側本数x2)

        rows.append({'ele': ele_no, 'sec_no': s_no, 'length': L, 't': t,
                     'D': D, 'concrete_m3': ele_concrete})

    return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
            'skipped': skipped}


# ---------------------------------------------------------------------------
# RCスラブ (板要素: PLATE/PLSTRS/PLSTRN/AXISYM)
# ---------------------------------------------------------------------------

def _rc_slab_elements(mgt_path):
    """RC(材料type3)のスラブ用板要素を厚みIDごとにまとめて返す.

    mgtopen_plate の8列目(iSUB)はPLATE要素でも小さい整数値を取り得て
    PLSTRS/PLSTRN/AXISYM(値3/4/5)と値だけでは区別できないため、
    ここでは板要素の種類を問わずRC材料のもの全てを対象とする
    (平面応力/平面変形/軸対称要素は通常の建築スラブでは稀)。

    戻り値: ({厚みID(int): {'area_m2': float, 'n_elements': int}}, [除外理由])
    """
    node = mgtopen_node(mgt_path)
    plate = np.asarray(mgtopen_plate(mgt_path), dtype=float)
    if not plate.size:
        return {}, {}
    plate = np.atleast_2d(plate)
    material = np.atleast_2d(mgtopen_material(mgt_path))
    pos = _node_pos(node)

    groups = {}
    skipped = []
    for r in range(plate.shape[0]):
        m_idx = find_index(material[:, 0], plate[r, 1])
        if m_idx == -1 or material[m_idx, 1] != 3:
            continue
        ele_no = int(plate[r, 0])
        thick_id = int(plate[r, 2])
        node_ids = [int(v) for v in plate[r, 3:7] if int(v) != 0]
        pts = [pos[n] for n in node_ids if n in pos]
        if len(pts) < 3:
            skipped.append({'ele': ele_no, 'sec_no': thick_id,
                            'reason': '節点が3点未満のため面積を計算できません'})
            continue
        area = _polygon_area(pts)
        g = groups.setdefault(thick_id, {'area_m2': 0.0, 'n_elements': 0})
        g['area_m2'] += area
        g['n_elements'] += 1
    return groups, skipped


def slab_plate_thicknesses(mgt_path):
    """RCスラブ用板要素が使う厚みID一覧を返す (配筋入力欄の候補用).

    戻り値: [{'thick_id','thickness_mm','name','n_elements','area_m2'}, ...]
    """
    groups, _skipped = _rc_slab_elements(mgt_path)
    if not groups:
        return []
    thickness = np.asarray(mgtopen_thickness(mgt_path), dtype=float)
    thickness_names = mgtopen_thickness_name(mgt_path)
    out = []
    for tid, g in sorted(groups.items()):
        mm = None
        if thickness.size:
            ti = find_index(np.atleast_2d(thickness)[:, 0], tid)
            if ti != -1:
                mm = float(np.atleast_2d(thickness)[ti, 1])
        out.append({'thick_id': tid, 'thickness_mm': mm,
                    'name': thickness_names.get(tid, ''),
                    'n_elements': g['n_elements'],
                    'area_m2': g['area_m2']})
    return out


def rc_slab_quantities(mgt_path, rebar_by_thickness):
    """RCスラブ(板要素)のX/Y方向筋質量とコンクリート体積.

    多数の小さなメッシュ要素で構成される連続体として扱い、部材(壁・柱)の
    ような要素単位の「+1」丸めは行わず、厚みID区分の合計面積をピッチで
    割って延長を概算する (壁・柱より粗い概算)。

    rebar_by_thickness: {厚みID(int): {'layer': 'single'(シングル配筋)|
                         'double'(ダブル配筋), 'top': {'x_dia','x_pitch',
                         'y_dia','y_pitch'}, 'bottom': {...}}}
      'bottom' はダブル配筋のときのみ使用し、上端筋・下端筋を別々に指定できる
      (シングル配筋のときは 'top' のみを1層として使う)。
    """
    groups, skipped = _rc_slab_elements(mgt_path)
    rebar = Tally()
    concrete_m3 = 0.0
    rows = []
    if not groups:
        return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
                'skipped': skipped}

    thickness = np.asarray(mgtopen_thickness(mgt_path), dtype=float)
    rebar_by_thickness = rebar_by_thickness or {}

    for tid, g in sorted(groups.items()):
        area = g['area_m2']
        t = 0.0
        if thickness.size:
            ti = find_index(np.atleast_2d(thickness)[:, 0], tid)
            if ti != -1:
                t = float(np.atleast_2d(thickness)[ti, 1]) / 1000.0
        ele_concrete = area * t
        concrete_m3 += ele_concrete

        cfg = rebar_by_thickness.get(tid)
        if cfg is None:
            rows.append({'thick_id': tid, 'n_elements': g['n_elements'],
                        'area_m2': area, 'thickness_m': t,
                        'concrete_m3': ele_concrete})
            continue

        layers = [cfg.get('top') or {}]
        if str(cfg.get('layer', 'double')) == 'double':
            layers.append(cfg.get('bottom') or {})
        for layer_cfg in layers:
            x_dia = float(layer_cfg.get('x_dia', 0) or 0)
            x_pitch = float(layer_cfg.get('x_pitch', 0) or 0)
            y_dia = float(layer_cfg.get('y_dia', 0) or 0)
            y_pitch = float(layer_cfg.get('y_pitch', 0) or 0)
            if x_dia > 0 and x_pitch > 0:
                rebar.add(x_dia, area / (x_pitch / 1000.0))
            if y_dia > 0 and y_pitch > 0:
                rebar.add(y_dia, area / (y_pitch / 1000.0))

        rows.append({'thick_id': tid, 'n_elements': g['n_elements'],
                     'area_m2': area, 'thickness_m': t,
                     'concrete_m3': ele_concrete})

    return {'rebar': rebar, 'concrete_m3': concrete_m3, 'rows': rows,
            'skipped': skipped}


# ---------------------------------------------------------------------------
# 統合・集計表
# ---------------------------------------------------------------------------

def compute_quantities(mgt_path, cover=None, wall_layer_by_thickness=None,
                       wall_line_rebar=None, slab_rebar_by_thickness=None):
    """RC梁・柱・壁(板要素壁+線材壁)・スラブ(板要素) の数量をまとめて計算する.

    cover: {'beam':mm,'column':mm,'wall':mm} 省略時は各40mm。
    wall_layer_by_thickness: {厚みID: 'single'|'double'} (rc_wall_quantities
      参照)。板要素壁 (*REBAR-WALL) の厚みごとの単配筋/複配筋指定
      (省略の厚みは複配筋)。
    wall_line_rebar: {断面番号: {...}} (rc_wall_line_quantities 参照)。
      断面名に'W'を含む線材壁パネルのユーザー入力配筋 (省略時はこの経路は
      対象なし)。結果は板要素壁と合算して 'wall' に含める
      ('wall_line_rows'/'wall_line_skipped' に線材壁のみの明細も残す)。
    slab_rebar_by_thickness: {厚みID: {...}} (rc_slab_quantities 参照)。
      板要素スラブの厚みごとのユーザー入力配筋 (省略時はこの経路は対象なし)。
    戻り値 dict: {'beam','column','wall','slab'} (各 rc_*_quantities の戻り値)
      + 'total_rebar' (Tally), 'total_concrete_m3' (float)
    """
    cover = cover or {}
    beam_cover = float(cover.get('beam', 40)) / 1000.0
    column_cover = float(cover.get('column', 40)) / 1000.0
    wall_cover = float(cover.get('wall', 40)) / 1000.0

    beam = rc_beam_quantities(mgt_path, beam_cover)
    column = rc_column_quantities(mgt_path, column_cover)
    wall = rc_wall_quantities(mgt_path, wall_cover,
                              layer_by_thickness=wall_layer_by_thickness)

    wall['line_rows'] = []
    if wall_line_rebar:
        wall_line = rc_wall_line_quantities(mgt_path, wall_line_rebar)
        wall['rebar'].merge(wall_line['rebar'])
        wall['concrete_m3'] += wall_line['concrete_m3']
        wall['line_rows'] = wall_line['rows']
        wall['skipped'] = wall['skipped'] + wall_line['skipped']

    slab = rc_slab_quantities(mgt_path, slab_rebar_by_thickness)

    total_rebar = Tally()
    total_rebar.merge(beam['rebar'])
    total_rebar.merge(column['rebar'])
    total_rebar.merge(wall['rebar'])
    total_rebar.merge(slab['rebar'])
    total_concrete = (beam['concrete_m3'] + column['concrete_m3']
                      + wall['concrete_m3'] + slab['concrete_m3'])

    return {'beam': beam, 'column': column, 'wall': wall, 'slab': slab,
            'total_rebar': total_rebar, 'total_concrete_m3': total_concrete}


def summarize_by_diameter(result):
    """径ごとの梁/柱/壁/スラブ/合計 質量(kg)・延長(m) 一覧 (径の昇順) を返す."""
    diameters = sorted(set(result['beam']['rebar']) | set(result['column']['rebar'])
                       | set(result['wall']['rebar']) | set(result['slab']['rebar']))
    out = []
    for d in diameters:
        b = result['beam']['rebar'].get(d, {'length': 0.0, 'weight': 0.0})
        c = result['column']['rebar'].get(d, {'length': 0.0, 'weight': 0.0})
        w = result['wall']['rebar'].get(d, {'length': 0.0, 'weight': 0.0})
        s = result['slab']['rebar'].get(d, {'length': 0.0, 'weight': 0.0})
        out.append({
            'dia': d,
            'beam_m': b['length'], 'beam_kg': b['weight'],
            'column_m': c['length'], 'column_kg': c['weight'],
            'wall_m': w['length'], 'wall_kg': w['weight'],
            'slab_m': s['length'], 'slab_kg': s['weight'],
            'total_m': b['length'] + c['length'] + w['length'] + s['length'],
            'total_kg': b['weight'] + c['weight'] + w['weight'] + s['weight'],
        })
    return out


def export_quantity_xlsx(result, out_path):
    """数量集計結果をExcel出力する (openpyxl).

    シート構成:
      径別集計   : 径ごとの梁/柱/壁/合計 質量(kg)・延長(m)
      コンクリート: 部材種別ごとの体積(m3)
      梁_明細/柱_明細/壁_明細: 部材別の明細行
    """
    import openpyxl

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = '径別集計'
    ws.append(['径', '梁(m)', '梁(kg)', '柱(m)', '柱(kg)', '壁(m)', '壁(kg)',
               'スラブ(m)', 'スラブ(kg)', '合計(m)', '合計(kg)'])
    for row in summarize_by_diameter(result):
        ws.append(['D%d' % row['dia'], row['beam_m'], row['beam_kg'],
                  row['column_m'], row['column_kg'], row['wall_m'],
                  row['wall_kg'], row['slab_m'], row['slab_kg'],
                  row['total_m'], row['total_kg']])
    total = result['total_rebar']
    ws.append(['合計', None, None, None, None, None, None, None, None,
              sum(v['length'] for v in total.values()),
              sum(v['weight'] for v in total.values())])

    ws2 = wb.create_sheet('コンクリート体積')
    ws2.append(['部材', '体積(m3)'])
    ws2.append(['梁', result['beam']['concrete_m3']])
    ws2.append(['柱', result['column']['concrete_m3']])
    ws2.append(['壁', result['wall']['concrete_m3']])
    ws2.append(['スラブ', result['slab']['concrete_m3']])
    ws2.append(['合計', result['total_concrete_m3']])

    ws3 = wb.create_sheet('梁_明細')
    ws3.append(['要素番号', '断面番号', '長さ(m)', 'b(m)', 'D(m)', '体積(m3)'])
    for r in result['beam']['rows']:
        ws3.append([r['ele'], r['sec_no'], r['length'], r['b'], r['D'],
                   r['concrete_m3']])

    ws4 = wb.create_sheet('柱_明細')
    ws4.append(['要素番号', '断面番号', '長さ(m)', 'b(m)', 'D(m)', '体積(m3)'])
    for r in result['column']['rows']:
        ws4.append([r['ele'], r['sec_no'], r['length'], r['b'], r['D'],
                   r['concrete_m3']])

    ws5 = wb.create_sheet('壁_明細(板要素)')
    ws5.append(['壁ID', 'ベースレベル', '厚みID', '幅(m)', '高さ(m)', '厚み(m)',
               '体積(m3)'])
    for r in result['wall']['rows']:
        ws5.append([r['wall_id'], r['base_level'], r.get('thick_id'),
                   r['width'], r['height'], r['thickness_m'],
                   r['concrete_m3']])

    if result['wall'].get('line_rows'):
        ws6 = wb.create_sheet('壁_明細(線材)')
        ws6.append(['要素番号', '断面番号', '高さ(m)', '厚みt(m)', '壁長D(m)',
                    '体積(m3)'])
        for r in result['wall']['line_rows']:
            ws6.append([r['ele'], r['sec_no'], r['length'], r['t'], r['D'],
                       r['concrete_m3']])

    ws7 = wb.create_sheet('スラブ_明細')
    ws7.append(['厚みID', '要素数', '面積(m2)', '厚み(m)', '体積(m3)'])
    for r in result['slab']['rows']:
        ws7.append([r['thick_id'], r['n_elements'], r['area_m2'],
                   r['thickness_m'], r['concrete_m3']])

    wb.save(out_path)
    return out_path
